"""
موتور Export به Excel با پشتیبانی Template

توسعه‌دهنده: علیرضا حامد
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple
import shutil

from app.core.logger import app_logger
from app.core.database import db_manager
from app.models import ExportTemplate, SalesData, ExportLog


class ExcelExporter:
    """کلاس Export داده‌ها به Excel"""
    
    def __init__(self):
        self.logger = app_logger
    
    def export_to_excel(
        self,
        template: ExportTemplate,
        data_list: List[SalesData],
        output_path: str
    ) -> Tuple[bool, str]:
        """
        Export داده‌ها به Excel بر اساس Template
        
        Args:
            template: Template Export
            data_list: لیست داده‌های استخراج شده
            output_path: مسیر فایل خروجی
            
        Returns:
            (موفقیت, پیام)
        """
        try:
            # بارگذاری فایل Template
            if not Path(template.template_path).exists():
                return False, f"فایل Template یافت نشد: {template.template_path}"
            
            # کپی Template به فایل خروجی
            shutil.copy2(template.template_path, output_path)
            
            # باز کردن فایل
            wb = openpyxl.load_workbook(output_path)
            
            # انتخاب Worksheet
            if template.target_worksheet in wb.sheetnames:
                ws = wb[template.target_worksheet]
            else:
                ws = wb.active
            
            # دریافت Mapping ستون‌ها
            column_mappings = template.column_mappings or {}
            
            # بررسی فرمت mapping (قدیم یا جدید)
            # فرمت قدیم: {'db_column': 'A'}
            # فرمت جدید: {'A': {'source_column': 'db_column', 'source_sheet': 1, 'formula': None}}
            is_new_format = False
            if column_mappings:
                first_value = next(iter(column_mappings.values()))
                if isinstance(first_value, dict) and 'source_column' in first_value:
                    is_new_format = True
            
            # تبدیل فرمت جدید به قدیم برای سازگاری
            if is_new_format:
                old_format_mappings = {}
                for excel_col, mapping_info in column_mappings.items():
                    source_col = mapping_info.get('source_column')
                    if source_col:
                        old_format_mappings[source_col] = excel_col
                column_mappings = old_format_mappings
            
            # شروع از سطر مشخص شده
            current_row = template.start_row
            
            # نوشتن داده‌ها
            for data in data_list:
                data_dict = data.data
                
                # پردازش هر ستون
                for db_column, excel_column in column_mappings.items():
                    if db_column in data_dict:
                        value = data_dict[db_column]
                        
                        # تعیین ستون Excel
                        if excel_column.isalpha():
                            # حرف ستون (A, B, C)
                            col = excel_column.upper()
                        else:
                            # اگر عدد است، تبدیل به حرف
                            try:
                                col = get_column_letter(int(excel_column))
                            except:
                                # اگر نام ستون است، از Mapping استفاده کن
                                col = excel_column
                        
                        # نوشتن مقدار
                        try:
                            cell = ws[f"{col}{current_row}"]
                            cell.value = value
                        except Exception as e:
                            self.logger.warning(f"خطا در نوشتن ستون {col}: {str(e)}")
                
                current_row += 1
            
            # ذخیره فایل
            wb.save(output_path)
            wb.close()
            
            # علامت‌گذاری داده‌ها به عنوان Exported
            data_ids = [data.id for data in data_list]
            db_manager.mark_as_exported(data_ids, template.template_type)
            
            # ثبت لاگ Export
            try:
                db = db_manager.get_session()
                export_log = ExportLog(
                    export_type=template.template_type,
                    record_count=len(data_list),
                    file_path=output_path,
                    template_name=template.name
                )
                db.add(export_log)
                db.commit()
                db.close()
                self.logger.info(f"لاگ Export ثبت شد (ID: {export_log.id})")
            except Exception as log_error:
                self.logger.warning(f"خطا در ثبت لاگ Export: {log_error}")
            
            message = f"✅ {len(data_list)} رکورد با موفقیت Export شدند"
            self.logger.success(message)
            
            return True, message
            
        except Exception as e:
            error_msg = f"خطا در Export: {str(e)}"
            self.logger.error(error_msg)
            return False, error_msg
    
    def export_with_formatting(
        self,
        template: ExportTemplate,
        data_list: List[SalesData],
        output_path: str,
        apply_styling: bool = True
    ) -> Tuple[bool, str]:
        """
        Export با فرمت‌بندی پیشرفته
        
        Args:
            template: Template Export
            data_list: لیست داده‌ها
            output_path: مسیر خروجی
            apply_styling: اعمال استایل‌ها
            
        Returns:
            (موفقیت, پیام)
        """
        try:
            # Export اولیه
            success, message = self.export_to_excel(template, data_list, output_path)
            
            if not success or not apply_styling:
                return success, message
            
            # اعمال فرمت‌بندی
            wb = openpyxl.load_workbook(output_path)
            ws = wb[template.target_worksheet] if template.target_worksheet in wb.sheetnames else wb.active
            
            # استایل‌های پیش‌فرض
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border_thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # تبدیل فرمت mapping (اگر جدید است)
            column_mappings = template.column_mappings or {}
            is_new_format = False
            if column_mappings:
                first_value = next(iter(column_mappings.values()))
                if isinstance(first_value, dict) and 'source_column' in first_value:
                    is_new_format = True
            
            if is_new_format:
                old_format_mappings = {}
                for excel_col, mapping_info in column_mappings.items():
                    source_col = mapping_info.get('source_column')
                    if source_col:
                        old_format_mappings[source_col] = excel_col
                column_mappings = old_format_mappings
            
            # فرمت‌بندی هدر (سطر قبل از start_row)
            if template.start_row > 1:
                header_row = template.start_row - 1
                for db_col, excel_col in column_mappings.items():
                    if isinstance(excel_col, str) and excel_col.isalpha():
                        cell = ws[f"{excel_col}{header_row}"]
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = border_thin
            
            # فرمت‌بندی داده‌ها
            end_row = template.start_row + len(data_list) - 1
            for row_idx in range(template.start_row, end_row + 1):
                for db_col, excel_col in column_mappings.items():
                    if isinstance(excel_col, str) and excel_col.isalpha():
                        cell = ws[f"{excel_col}{row_idx}"]
                        cell.border = border_thin
                        cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # تنظیم عرض ستون‌ها خودکار
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # ذخیره
            wb.save(output_path)
            wb.close()
            
            return True, f"✅ Export با فرمت‌بندی انجام شد: {len(data_list)} رکورد"
            
        except Exception as e:
            error_msg = f"خطا در فرمت‌بندی: {str(e)}"
            self.logger.error(error_msg)
            return False, error_msg
    
    def generate_output_filename(self, template: ExportTemplate, sheet_name: str = "") -> str:
        """
        ساخت نام فایل خروجی
        
        Args:
            template: Template Export
            sheet_name: نام شیت (اختیاری)
            
        Returns:
            نام فایل
        """
        pattern = template.output_filename_pattern or "Export_{type}_{date}.xlsx"
        
        replacements = {
            '{type}': template.template_type,
            '{name}': template.name.replace(' ', '_'),
            '{sheet}': sheet_name.replace(' ', '_') if sheet_name else '',
            '{date}': datetime.now().strftime('%Y-%m-%d'),
            '{datetime}': datetime.now().strftime('%Y-%m-%d_%H-%M-%S'),
            '{time}': datetime.now().strftime('%H-%M-%S')
        }
        
        filename = pattern
        for key, value in replacements.items():
            filename = filename.replace(key, value)
        
        return filename


# نمونه سراسری
excel_exporter = ExcelExporter()
