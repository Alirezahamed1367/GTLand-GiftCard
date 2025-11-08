"""
Helper برای خواندن اطلاعات از فایل‌های Excel
"""
import openpyxl
from typing import List, Dict, Optional
from pathlib import Path


class ExcelColumnInfo:
    """اطلاعات یک ستون Excel"""
    def __init__(self, letter: str, index: int, name: str = None):
        self.letter = letter  # A, B, C, ...
        self.index = index    # 1, 2, 3, ...
        self.name = name      # نام ستون (اگر در ردیف اول باشد)
    
    def __repr__(self):
        if self.name:
            return f"{self.letter} ({self.name})"
        return self.letter
    
    def to_dict(self):
        return {
            'letter': self.letter,
            'index': self.index,
            'name': self.name
        }


class ExcelHelper:
    """کلاس کمکی برای کار با Excel"""
    
    @staticmethod
    def get_worksheets(file_path: str) -> List[str]:
        """
        دریافت لیست Worksheet های موجود در فایل Excel
        
        Args:
            file_path: مسیر فایل Excel
            
        Returns:
            لیست نام Worksheet ها
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            worksheets = workbook.sheetnames
            workbook.close()
            return worksheets
        except Exception as e:
            print(f"خطا در خواندن Worksheet ها: {str(e)}")
            return []
    
    @staticmethod
    def get_columns(file_path: str, worksheet_name: str = None, header_row: int = 1) -> List[ExcelColumnInfo]:
        """
        دریافت اطلاعات ستون‌های یک Worksheet
        
        Args:
            file_path: مسیر فایل Excel
            worksheet_name: نام Worksheet (None برای اولین Sheet)
            header_row: شماره ردیف هدر (پیش‌فرض: 1)
            
        Returns:
            لیست اطلاعات ستون‌ها
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # انتخاب Worksheet
            if worksheet_name:
                worksheet = workbook[worksheet_name]
            else:
                worksheet = workbook.active
            
            columns = []
            
            # خواندن ردیف هدر
            for cell in worksheet[header_row]:
                if cell.value is not None or cell.column <= 50:  # حداکثر 50 ستون
                    col_info = ExcelColumnInfo(
                        letter=cell.column_letter,
                        index=cell.column,
                        name=str(cell.value) if cell.value is not None else None
                    )
                    columns.append(col_info)
                
                # متوقف شدن در اولین ستون خالی (بعد از 3 ستون خالی متوالی)
                if cell.value is None and cell.column > 3:
                    # بررسی 2 ستون بعدی
                    next_cells = list(worksheet[header_row])[cell.column:cell.column+2]
                    if all(c.value is None for c in next_cells):
                        break
            
            workbook.close()
            return columns
            
        except Exception as e:
            print(f"خطا در خواندن ستون‌ها: {str(e)}")
            return []
    
    @staticmethod
    def get_column_data(file_path: str, worksheet_name: str, column_letter: str, 
                       start_row: int = 2, max_rows: int = 10) -> List:
        """
        دریافت نمونه داده‌های یک ستون (برای پیش‌نمایش)
        
        Args:
            file_path: مسیر فایل Excel
            worksheet_name: نام Worksheet
            column_letter: حرف ستون (A, B, C, ...)
            start_row: ردیف شروع
            max_rows: حداکثر تعداد ردیف
            
        Returns:
            لیست داده‌ها
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            worksheet = workbook[worksheet_name]
            
            data = []
            for row in range(start_row, start_row + max_rows):
                cell = worksheet[f"{column_letter}{row}"]
                data.append(cell.value)
            
            workbook.close()
            return data
            
        except Exception as e:
            print(f"خطا در خواندن داده‌ها: {str(e)}")
            return []
    
    @staticmethod
    def validate_file(file_path: str) -> tuple[bool, str]:
        """
        اعتبارسنجی فایل Excel
        
        Args:
            file_path: مسیر فایل
            
        Returns:
            (موفق, پیام)
        """
        path = Path(file_path)
        
        # بررسی وجود فایل
        if not path.exists():
            return False, "فایل یافت نشد!"
        
        # بررسی پسوند
        if path.suffix.lower() not in ['.xlsx', '.xls', '.xlsm']:
            return False, "فرمت فایل باید Excel باشد (.xlsx, .xls, .xlsm)"
        
        # تلاش برای باز کردن
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            workbook.close()
            return True, "فایل معتبر است"
        except Exception as e:
            return False, f"خطا در باز کردن فایل: {str(e)}"
    
    @staticmethod
    def get_file_info(file_path: str) -> Dict:
        """
        دریافت اطلاعات کلی فایل Excel
        
        Args:
            file_path: مسیر فایل
            
        Returns:
            دیکشنری اطلاعات
        """
        try:
            path = Path(file_path)
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            info = {
                'filename': path.name,
                'size': path.stat().st_size,
                'size_mb': round(path.stat().st_size / 1024 / 1024, 2),
                'worksheets': workbook.sheetnames,
                'worksheet_count': len(workbook.sheetnames),
                'active_sheet': workbook.active.title
            }
            
            workbook.close()
            return info
            
        except Exception as e:
            return {
                'error': str(e)
            }


# توابع کمکی سریع
def get_excel_columns(file_path: str, worksheet_name: str = None) -> List[Dict]:
    """
    تابع سریع برای دریافت ستون‌ها به صورت دیکشنری
    
    Returns:
        لیست دیکشنری ستون‌ها
    """
    helper = ExcelHelper()
    columns = helper.get_columns(file_path, worksheet_name)
    return [col.to_dict() for col in columns]


def get_excel_worksheets(file_path: str) -> List[str]:
    """تابع سریع برای دریافت Worksheet ها"""
    helper = ExcelHelper()
    return helper.get_worksheets(file_path)


if __name__ == "__main__":
    # تست
    print("🧪 تست ExcelHelper:")
    
    # مثال
    test_file = "templates/sample.xlsx"
    
    if Path(test_file).exists():
        print(f"\n📄 فایل: {test_file}")
        
        # Worksheet ها
        sheets = get_excel_worksheets(test_file)
        print(f"\n📊 Worksheet ها: {sheets}")
        
        # ستون‌ها
        if sheets:
            columns = get_excel_columns(test_file, sheets[0])
            print(f"\n📋 ستون‌ها در {sheets[0]}:")
            for col in columns:
                print(f"  {col['letter']}: {col['name'] or '(بدون نام)'}")
    else:
        print(f"⚠️ فایل تست یافت نشد: {test_file}")
