"""
Dialog مدیریت پیشرفته Template های Export با Mapping و Formula
"""
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QMessageBox, QWizard, QWizardPage,
    QFileDialog, QLineEdit, QComboBox, QTextEdit,
    QGroupBox, QListWidget, QWidget
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QScreen
import json
from pathlib import Path
import openpyxl

from app.core.database import db_manager
from app.core.logger import app_logger
from app.models import ExportTemplate
from app.utils.ui_constants import (
    FONT_SIZE_TITLE, FONT_SIZE_LABEL, FONT_SIZE_NORMAL,
    BUTTON_HEIGHT_LARGE, BUTTON_HEIGHT_MEDIUM,
    COLOR_PRIMARY, COLOR_SUCCESS, COLOR_DANGER,
    get_button_style
)
from app.gui.widgets.column_mapping_widget import ColumnMappingWidget


class TemplateWizard(QWizard):
    """
    ویزارد ساخت Template جدید
    
    مراحل:
    1. انتخاب فایل Excel نمونه
    2. تحلیل فایل و نمایش ستون‌ها
    3. انتخاب Google Sheets مبدا
    4. Mapping ستون‌ها با Drag & Drop
    5. تنظیمات و ذخیره
    """
    
    PAGE_SELECT_FILE = 0
    PAGE_ANALYZE_FILE = 1
    PAGE_SELECT_SHEETS = 2
    PAGE_MAPPING = 3
    PAGE_SETTINGS = 4
    
    def __init__(self, parent=None, template=None):
        super().__init__(parent)
        self.template = template  # برای ویرایش
        self.excel_file_path = None
        self.excel_columns = []
        self.available_sheets = {}
        self.column_mappings = {}
        
        self.setWindowTitle("⚡ ساخت Template پیشرفته" if not template else "⚡ ویرایش Template")
        self.setWizardStyle(QWizard.WizardStyle.ModernStyle)
        self.setMinimumSize(1000, 700)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        # افزودن صفحات
        self.setPage(self.PAGE_SELECT_FILE, self.create_select_file_page())
        self.setPage(self.PAGE_ANALYZE_FILE, self.create_analyze_file_page())
        self.setPage(self.PAGE_SELECT_SHEETS, self.create_select_sheets_page())
        self.setPage(self.PAGE_MAPPING, self.create_mapping_page())
        self.setPage(self.PAGE_SETTINGS, self.create_settings_page())
        
        # شروع از اولین صفحه
        self.setStartId(self.PAGE_SELECT_FILE)
        
        # اگر در حال ویرایش است، داده‌های Template را بارگذاری کن
        if self.template:
            self.load_template_data()
        
        # استایل یکپارچه با سایر فرم‌ها
        self.setStyleSheet(f"""
            QWizard {{
                background-color: #f8f9fa;
            }}
            QWizardPage {{
                background-color: white;
            }}
            QGroupBox {{
                font-weight: bold;
                font-size: {FONT_SIZE_LABEL}pt;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 8px;
                background-color: #fafafa;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                background-color: white;
            }}
            QLabel {{
                font-size: {FONT_SIZE_NORMAL}pt;
            }}
            QLineEdit, QTextEdit, QComboBox {{
                border: 2px solid #e0e0e0;
                border-radius: 5px;
                padding: 8px;
                font-size: {FONT_SIZE_NORMAL}pt;
                background-color: white;
            }}
            QLineEdit:focus, QTextEdit:focus, QComboBox:focus {{
                border: 2px solid {COLOR_PRIMARY};
            }}
            QPushButton {{
                border-radius: 8px;
                font-weight: bold;
                padding: 10px;
            }}
        """)
    
    def create_select_file_page(self):
        """صفحه 1: انتخاب فایل Excel"""
        page = QWizardPage()
        page.setTitle("📄 انتخاب فایل Excel نمونه")
        page.setSubTitle("فایل Excel که از نرم‌افزار مالیاتی خود دریافت کرده‌اید را انتخاب کنید")
        
        layout = QVBoxLayout(page)
        layout.setSpacing(20)
        
        # توضیحات
        info = QLabel("""
        <div style='background: #E3F2FD; padding: 15px; border-radius: 8px;'>
            <h3>📋 این فایل چیست؟</h3>
            <p>فایل Excel نمونه‌ای که از نرم‌افزار مالیاتی شما (مثل سامانه مؤدیان) دریافت کرده‌اید.</p>
            <p>این فایل شامل ساختار و عناوین ستون‌هایی است که داده‌ها باید در آن قرار بگیرند.</p>
            <br>
            <h4>✅ نکات مهم:</h4>
            <ul>
                <li>فایل باید فرمت .xlsx یا .xls باشد</li>
                <li>ردیف اول باید حاوی عناوین ستون‌ها باشد</li>
                <li>می‌توانید چند Sheet در فایل داشته باشید</li>
            </ul>
        </div>
        """)
        info.setWordWrap(True)
        layout.addWidget(info)
        
        # دکمه انتخاب فایل
        file_group = QGroupBox("📂 فایل Excel")
        file_layout = QVBoxLayout(file_group)
        
        self.file_path_label = QLabel("هیچ فایلی انتخاب نشده")
        self.file_path_label.setStyleSheet("padding: 10px; background: #f5f5f5; border-radius: 5px;")
        file_layout.addWidget(self.file_path_label)
        
        select_file_btn = QPushButton("📁 انتخاب فایل Excel")
        select_file_btn.setMinimumHeight(BUTTON_HEIGHT_LARGE)
        select_file_btn.setStyleSheet(get_button_style(COLOR_PRIMARY, FONT_SIZE_LABEL, BUTTON_HEIGHT_LARGE))
        select_file_btn.clicked.connect(self.select_excel_file)
        file_layout.addWidget(select_file_btn)
        
        layout.addWidget(file_group)
        layout.addStretch()
        
        return page
    
    def create_analyze_file_page(self):
        """صفحه 2: تحلیل فایل"""
        page = QWizardPage()
        page.setTitle("🔍 تحلیل فایل Excel")
        page.setSubTitle("اطلاعات فایل انتخابی")
        
        layout = QVBoxLayout(page)
        
        self.analysis_text = QTextEdit()
        self.analysis_text.setReadOnly(True)
        self.analysis_text.setStyleSheet("""
            QTextEdit {
                font-family: 'Courier New';
                font-size: 11pt;
                background: #F5F5F5;
                border: 2px solid #2196F3;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.analysis_text)
        
        return page
    
    def create_select_sheets_page(self):
        """صفحه 3: انتخاب Google Sheets"""
        page = QWizardPage()
        page.setTitle("📊 انتخاب Google Sheets مبدا")
        page.setSubTitle("Google Sheet هایی که داده از آن‌ها استخراج می‌شود را انتخاب کنید")
        
        layout = QVBoxLayout(page)
        
        info = QLabel("""
        <div style='background: #C8E6C9; padding: 15px; border-radius: 8px;'>
            <h4>💡 راهنما:</h4>
            <p>داده‌های شما از Google Sheets مختلف استخراج می‌شوند.</p>
            <p>در این مرحله، Sheet هایی که می‌خواهید از آن‌ها داده بگیرید را مشخص کنید.</p>
        </div>
        """)
        info.setWordWrap(True)
        layout.addWidget(info)
        
        # لیست Sheet های موجود
        sheets_group = QGroupBox("📄 Google Sheets موجود")
        sheets_layout = QVBoxLayout(sheets_group)
        
        self.sheets_list = QListWidget()
        self.sheets_list.setStyleSheet("""
            QListWidget {
                font-size: 11pt;
                padding: 5px;
            }
            QListWidget::item {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                margin: 3px;
            }
            QListWidget::item:selected {
                background: #4CAF50;
                color: white;
            }
        """)
        self.sheets_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        sheets_layout.addWidget(self.sheets_list)
        
        layout.addWidget(sheets_group)
        
        # بارگذاری Sheet ها
        self.load_available_sheets()
        
        return page
    
    def create_mapping_page(self):
        """صفحه 4: Mapping ستون‌ها"""
        page = QWizardPage()
        page.setTitle("🔗 Mapping ستون‌ها")
        page.setSubTitle("ستون‌های Google Sheet را به ستون‌های Excel متصل کنید")
        
        layout = QVBoxLayout(page)
        
        # ویجت Mapping (خواهد شد زمانی که صفحه باز شود)
        self.mapping_widget = None
        self.mapping_widget_container = layout
        
        return page
    
    def create_settings_page(self):
        """صفحه 5: تنظیمات نهایی"""
        page = QWizardPage()
        page.setTitle("⚙️ تنظیمات Template")
        page.setSubTitle("نام و توضیحات Template را وارد کنید")
        
        layout = QVBoxLayout(page)
        layout.setSpacing(15)
        
        # نام Template
        name_group = QGroupBox("📝 نام Template")
        name_layout = QVBoxLayout(name_group)
        
        self.template_name = QLineEdit()
        self.template_name.setPlaceholderText("مثال: Template سامانه مؤدیان 1402")
        self.template_name.setMinimumHeight(40)
        self.template_name.setStyleSheet("font-size: 11pt; padding: 8px;")
        name_layout.addWidget(self.template_name)
        
        layout.addWidget(name_group)
        
        # نوع Template
        type_group = QGroupBox("🏷️ نوع Template")
        type_layout = QVBoxLayout(type_group)
        
        # توضیحات
        type_info = QLabel("💡 می‌توانید یکی از انواع از پیش تعریف شده را انتخاب کنید، یا نوع Custom با نام دلخواه:")
        type_info.setWordWrap(True)
        type_info.setStyleSheet("color: #666; font-size: 10pt; padding: 5px; background: #FFF9C4; border-radius: 4px;")
        type_layout.addWidget(type_info)
        
        # لیست کشویی برای انتخاب نوع
        type_select_layout = QHBoxLayout()
        type_select_layout.addWidget(QLabel("انتخاب نوع:"))
        
        self.template_type = QComboBox()
        self.template_type.addItems([
            "💰 حساب",
            "📄 فاکتور",
            "📦 کالا",
            "⚙️ سفارشی"
        ])
        self.template_type.setMinimumHeight(40)
        self.template_type.setStyleSheet("font-size: 11pt; padding: 8px;")
        self.template_type.currentIndexChanged.connect(self.on_template_type_changed)
        type_select_layout.addWidget(self.template_type, 1)
        type_layout.addLayout(type_select_layout)
        
        # فیلد نام Custom (مخفی به صورت پیش‌فرض)
        self.custom_type_widget = QWidget()
        custom_type_layout = QVBoxLayout(self.custom_type_widget)
        custom_type_layout.setContentsMargins(0, 5, 0, 0)
        
        custom_label = QLabel("📝 نام Custom:")
        custom_label.setStyleSheet("font-weight: bold;")
        custom_type_layout.addWidget(custom_label)
        
        self.custom_type_name = QLineEdit()
        self.custom_type_name.setPlaceholderText("مثال: تایپ ویژه شرکت ABC")
        self.custom_type_name.setMinimumHeight(40)
        self.custom_type_name.setStyleSheet("font-size: 11pt; padding: 8px; border: 2px solid #FF9800;")
        custom_type_layout.addWidget(self.custom_type_name)
        
        self.custom_type_widget.setVisible(False)
        type_layout.addWidget(self.custom_type_widget)
        
        layout.addWidget(type_group)
        
        # توضیحات
        desc_group = QGroupBox("📄 توضیحات")
        desc_layout = QVBoxLayout(desc_group)
        
        self.template_description = QTextEdit()
        self.template_description.setPlaceholderText("توضیحات اختیاری...")
        self.template_description.setMaximumHeight(100)
        self.template_description.setStyleSheet("font-size: 10pt; padding: 5px;")
        desc_layout.addWidget(self.template_description)
        
        layout.addWidget(desc_group)
        
        # خلاصه
        summary_group = QGroupBox("📊 خلاصه Template")
        summary_layout = QVBoxLayout(summary_group)
        
        self.summary_label = QLabel()
        self.summary_label.setWordWrap(True)
        self.summary_label.setStyleSheet("background: #F5F5F5; padding: 10px; border-radius: 5px;")
        summary_layout.addWidget(self.summary_label)
        
        layout.addWidget(summary_group)
        layout.addStretch()
        
        return page
    
    def select_excel_file(self):
        """انتخاب فایل Excel"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "انتخاب فایل Excel نمونه",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.excel_file_path = file_path
            self.file_path_label.setText(f"✅ {Path(file_path).name}")
            self.file_path_label.setStyleSheet("""
                padding: 10px;
                background: #C8E6C9;
                border: 2px solid #4CAF50;
                border-radius: 5px;
                font-weight: bold;
            """)
            
            # تحلیل فایل
            self.analyze_excel_file()
    
    def analyze_excel_file(self):
        """تحلیل فایل Excel"""
        if not self.excel_file_path:
            return
        
        try:
            workbook = openpyxl.load_workbook(self.excel_file_path, read_only=True)
            
            analysis = "📊 **تحلیل فایل Excel:**\n\n"
            analysis += f"📁 نام فایل: {Path(self.excel_file_path).name}\n"
            analysis += f"📄 تعداد Sheet ها: {len(workbook.sheetnames)}\n\n"
            
            # تحلیل هر Sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                analysis += f"{'='*50}\n"
                analysis += f"📋 Sheet: {sheet_name}\n"
                analysis += f"{'='*50}\n"
                
                # خواندن ردیف اول (عناوین)
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first_row:
                    columns = [col for col in first_row if col]
                    analysis += f"📊 تعداد ستون‌ها: {len(columns)}\n"
                    analysis += f"📝 عناوین ستون‌ها:\n"
                    
                    for i, col in enumerate(columns, 1):
                        col_letter = openpyxl.utils.get_column_letter(i)
                        analysis += f"   {col_letter}: {col}\n"
                        
                        # ذخیره برای صفحه بعد
                        if sheet_name == workbook.sheetnames[0]:  # فقط Sheet اول
                            self.excel_columns.append({
                                'letter': col_letter,
                                'name': str(col),
                                'sheet': sheet_name
                            })
                
                analysis += "\n"
            
            workbook.close()
            
            self.analysis_text.setText(analysis)
            
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در تحلیل فایل:\n{str(e)}")
            app_logger.error(f"Error analyzing Excel file: {e}")
    
    def load_template_data(self):
        """بارگذاری داده‌های Template برای ویرایش"""
        try:
            if not self.template:
                return
            
            # بارگذاری مسیر Excel
            self.excel_file_path = self.template.template_path
            if self.excel_file_path and Path(self.excel_file_path).exists():
                self.file_path_label.setText(f"✅ {Path(self.excel_file_path).name}")
                self.file_path_label.setStyleSheet("""
                    padding: 10px;
                    background: #C8E6C9;
                    border: 2px solid #4CAF50;
                    border-radius: 5px;
                    font-weight: bold;
                """)
                
                # تحلیل فایل Excel
                self.analyze_excel_file()
            
            # بارگذاری mappings
            if self.template.column_mappings:
                self.column_mappings = self.template.column_mappings
            
            # بارگذاری نام و توضیحات
            if hasattr(self, 'template_name'):
                self.template_name.setText(self.template.name)
            
            if hasattr(self, 'template_description'):
                self.template_description.setText(self.template.description or "")
            
            # بارگذاری نوع Template
            if hasattr(self, 'template_type'):
                template_type = self.template.template_type
                # پیدا کردن index مناسب
                for i in range(self.template_type.count()):
                    item_text = self.template_type.itemText(i)
                    if template_type.startswith("Custom:"):
                        # اگر Custom است
                        if "Custom" in item_text:
                            self.template_type.setCurrentIndex(i)
                            custom_name = template_type.replace("Custom: ", "").strip()
                            if hasattr(self, 'custom_type_name'):
                                self.custom_type_name.setText(custom_name)
                            break
                    elif template_type in item_text:
                        self.template_type.setCurrentIndex(i)
                        break
            
            app_logger.info(f"Template '{self.template.name}' بارگذاری شد برای ویرایش")
            
        except Exception as e:
            app_logger.error(f"خطا در بارگذاری Template: {e}")
            QMessageBox.warning(self, "هشدار", f"خطا در بارگذاری Template:\n{str(e)}")
    
    def load_available_sheets(self):
        """بارگذاری Google Sheets موجود"""
        try:
            self.sheets_list.clear()
            
            # دریافت از دیتابیس
            sheet_configs = db_manager.get_all_sheet_configs(active_only=True)
            
            if not sheet_configs:
                # اگر هیچ Sheet ای نیست، یک پیام نمایش بده
                self.sheets_list.addItem("⚠️ هیچ Google Sheet فعالی یافت نشد")
                return
            
            # Import Google Sheets برای خواندن ستون‌های واقعی
            from app.core.google_sheets import GoogleSheetExtractor
            extractor = GoogleSheetExtractor()
            
            for config in sheet_configs:
                # تحلیل ستون‌ها
                columns = []
                
                # ابتدا سعی می‌کنیم از Google Sheets واقعی بخوانیم
                try:
                    if config.sheet_url and config.worksheet_name:
                        headers = extractor.get_headers(config.sheet_url, config.worksheet_name)
                        if headers:
                            columns = headers
                            app_logger.info(f"✅ ستون‌های Sheet '{config.name}' از Google خوانده شد: {len(columns)} ستون")
                except Exception as e:
                    app_logger.warning(f"⚠️ خطا در خواندن ستون‌های واقعی از Google Sheets: {e}")
                
                # اگر نتوانستیم بخوانیم، از column_mappings استفاده کن
                if not columns and config.column_mappings:
                    columns = list(config.column_mappings.keys())
                    app_logger.info(f"📋 از column_mappings استفاده شد: {len(columns)} ستون")
                
                # اگر هنوز خالی است، پیام خطا
                if not columns:
                    app_logger.error(f"❌ Sheet '{config.name}' هیچ ستونی ندارد!")
                    continue
                
                self.available_sheets[config.id] = {
                    'name': config.name,
                    'worksheet': config.worksheet_name,
                    'sheet_url': config.sheet_url,
                    'columns': columns
                }
                
                # افزودن به لیست
                item_text = f"📊 {config.name}"
                if config.worksheet_name:
                    item_text += f" → {config.worksheet_name}"
                item_text += f" ({len(columns)} ستون)"
                
                self.sheets_list.addItem(item_text)
                self.sheets_list.item(self.sheets_list.count() - 1).setData(Qt.ItemDataRole.UserRole, config.id)
            
            # انتخاب همه به صورت پیش‌فرض
            for i in range(self.sheets_list.count()):
                self.sheets_list.item(i).setSelected(True)
        
        except Exception as e:
            app_logger.error(f"Error loading sheet configs: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "خطا", f"خطا در بارگذاری Google Sheets:\n{str(e)}")
    
    def initializePage(self, page_id):
        """هنگام ورود به هر صفحه"""
        try:
            if page_id == self.PAGE_MAPPING:
                # ساخت ویجت Mapping
                if not self.mapping_widget:
                    # دریافت Sheet های انتخابی
                    selected_sheets = {}
                    for item in self.sheets_list.selectedItems():
                        sheet_id = item.data(Qt.ItemDataRole.UserRole)
                        if sheet_id and sheet_id in self.available_sheets:
                            selected_sheets[sheet_id] = self.available_sheets[sheet_id]
                    
                    # بررسی اینکه حداقل یک Sheet داریم
                    if not selected_sheets:
                        QMessageBox.warning(
                            self,
                            "هشدار",
                            "هیچ Google Sheet فعالی انتخاب نشده است.\n\n"
                            "لطفاً ابتدا از منوی اصلی یک Google Sheet تنظیم کنید."
                        )
                        return
                    
                    # بررسی اینکه Excel columns داریم
                    if not self.excel_columns:
                        QMessageBox.warning(
                            self,
                            "هشدار",
                            "ستون‌های Excel شناسایی نشدند.\n\n"
                            "لطفاً یک فایل Excel معتبر انتخاب کنید."
                        )
                        return
                    
                    # ساخت ویجت
                    # پاس دادن کامل اطلاعات ستون‌ها (letter + name)
                    self.mapping_widget = ColumnMappingWidget(
                        self,
                        excel_columns=self.excel_columns,  # کامل: [{'letter': 'A', 'name': 'ردیف', 'sheet': ...}]
                        available_sheets=selected_sheets
                    )
                    self.mapping_widget_container.addWidget(self.mapping_widget)
            
            elif page_id == self.PAGE_SETTINGS:
                # به‌روزرسانی خلاصه
                self.update_summary()
        
        except Exception as e:
            app_logger.error(f"Error in initializePage: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "خطا", f"خطا در بارگذاری صفحه:\n{str(e)}")
    
    def update_summary(self):
        """به‌روزرسانی خلاصه Template"""
        if self.mapping_widget:
            mappings = self.mapping_widget.get_mappings()
            
            summary = f"""
            <h4>📋 خلاصه Template:</h4>
            <p><b>📁 فایل Excel:</b> {Path(self.excel_file_path).name if self.excel_file_path else '-'}</p>
            <p><b>📊 تعداد ستون‌های Excel:</b> {len(self.excel_columns)}</p>
            <p><b>✅ تعداد Mapping ها:</b> {len(mappings)}</p>
            <p><b>⚡ تعداد Formula ها:</b> {sum(1 for m in mappings.values() if m.get('formula'))}</p>
            """
            
            self.summary_label.setText(summary)
    
    def on_template_type_changed(self, index):
        """تغییر نوع Template"""
        # اگر Custom انتخاب شد (آخرین گزینه)، فیلد نام Custom را نمایش بده
        is_custom = (index == self.template_type.count() - 1)
        self.custom_type_widget.setVisible(is_custom)
    
    def validateCurrentPage(self):
        """اعتبارسنجی صفحه فعلی"""
        current_id = self.currentId()
        
        if current_id == self.PAGE_SELECT_FILE:
            if not self.excel_file_path:
                QMessageBox.warning(self, "هشدار", "لطفاً یک فایل Excel انتخاب کنید")
                return False
        
        elif current_id == self.PAGE_SELECT_SHEETS:
            if not self.sheets_list.selectedItems():
                QMessageBox.warning(self, "هشدار", "لطفاً حداقل یک Google Sheet انتخاب کنید")
                return False
        
        elif current_id == self.PAGE_SETTINGS:
            if not self.template_name.text().strip():
                QMessageBox.warning(self, "هشدار", "لطفاً نام Template را وارد کنید")
                return False
            
            # بررسی Custom Type
            if self.template_type.currentIndex() == self.template_type.count() - 1:
                if not self.custom_type_name.text().strip():
                    QMessageBox.warning(self, "هشدار", "لطفاً نام Custom برای نوع Template را وارد کنید")
                    return False
        
        return True
    
    def accept(self):
        """ذخیره Template"""
        try:
            # ✅ اعتبارسنجی قبل از ذخیره
            template_name = self.template_name.text().strip()
            if not template_name:
                QMessageBox.warning(self, "هشدار", "لطفاً نام Template را وارد کنید")
                return  # خروج بدون ذخیره
            
            # دریافت اطلاعات
            mappings = self.mapping_widget.get_mappings() if self.mapping_widget else {}
            
            # تعیین نوع Template
            template_type_text = self.template_type.currentText()
            if "Custom" in template_type_text:
                # اگر Custom است، از نام وارد شده استفاده کن
                custom_name = self.custom_type_name.text().strip()
                if not custom_name:
                    QMessageBox.warning(self, "هشدار", "لطفاً نام Custom برای نوع Template را وارد کنید")
                    return  # خروج بدون ذخیره
                final_type = f"Custom: {custom_name}"
            else:
                # استخراج نام کوتاه (مثلاً "حساب" از "💰 حساب")
                # حذف emoji و فاصله اضافی
                final_type = template_type_text.split(" ")[-1].strip()
            
            # اعتبارسنجی mappings
            if not mappings:
                reply = QMessageBox.question(
                    self,
                    "هشدار",
                    "هیچ Mapping ای تعریف نشده است.\n\nآیا مطمئن هستید که می‌خواهید ادامه دهید؟",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.No:
                    return  # خروج بدون ذخیره
            
            template_data = {
                'name': template_name,
                'template_type': final_type,
                'template_path': self.excel_file_path,
                'target_worksheet': self.excel_columns[0]['sheet'] if self.excel_columns else 'Sheet1',
                'column_mappings': mappings,
                'start_row': 2,
                'start_column': 1,
                'is_active': True,
                'description': self.template_description.toPlainText().strip()
            }
            
            # ذخیره در دیتابیس
            if self.template:
                # ویرایش
                success, message = db_manager.update_template(self.template.id, template_data)
            else:
                # جدید
                success, template, message = db_manager.create_template(template_data)
            
            if success:
                QMessageBox.information(self, "موفقیت", "Template با موفقیت ذخیره شد ✅")
                super().accept()  # فقط در صورت موفقیت
            else:
                QMessageBox.critical(self, "خطا", f"خطا در ذخیره Template:\n{message}")
                # در صورت خطا، باز نمی‌گردد
        
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در ذخیره Template:\n{str(e)}")
            app_logger.error(f"Error saving template: {e}")
            # در صورت خطا، باز نمی‌گردد


class TemplateManagerDialog(QDialog):
    """Dialog مدیریت Template های Export"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()
        self.load_templates()
    
    def init_ui(self):
        """راه‌اندازی رابط کاربری"""
        self.setWindowTitle("مدیریت Template ها")
        self.resize(900, 600)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        # استایل یکپارچه با سایر دیالوگ‌ها
        self.setStyleSheet(f"""
            QDialog {{
                background-color: #f8f9fa;
            }}
            QLabel {{
                font-size: {FONT_SIZE_NORMAL}pt;
            }}
            QGroupBox {{
                font-weight: bold;
                font-size: {FONT_SIZE_LABEL}pt;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 8px;
                background-color: white;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                background-color: #f8f9fa;
            }}
            QListWidget {{
                background-color: white;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                padding: 5px;
            }}
            QListWidget::item {{
                padding: 12px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                margin: 5px;
                background: white;
            }}
            QListWidget::item:selected {{
                background: #E3F2FD;
                border: 2px solid {COLOR_PRIMARY};
            }}
            QListWidget::item:hover {{
                background: #f5f5f5;
            }}
        """)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # عنوان
        title = QLabel("📋 مدیریت Template های Export")
        title_font = QFont()
        title_font.setPointSize(FONT_SIZE_TITLE)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet(f"color: {COLOR_PRIMARY}; padding: 10px;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        # لیست Template ها
        self.templates_list = QListWidget()
        self.templates_list.setStyleSheet(f"""
            QListWidget {{
                font-size: {FONT_SIZE_NORMAL}pt;
            }}
        """)
        layout.addWidget(self.templates_list, 1)
        
        # دکمه‌ها
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)
        
        add_btn = QPushButton("➕ ساخت Template جدید")
        add_btn.setMinimumHeight(BUTTON_HEIGHT_LARGE)
        add_btn.setStyleSheet(get_button_style(COLOR_SUCCESS, FONT_SIZE_LABEL, BUTTON_HEIGHT_LARGE))
        add_btn.clicked.connect(self.add_template)
        buttons_layout.addWidget(add_btn)
        
        edit_btn = QPushButton("✏️ ویرایش")
        edit_btn.setMinimumHeight(BUTTON_HEIGHT_LARGE)
        edit_btn.setStyleSheet(get_button_style(COLOR_PRIMARY, FONT_SIZE_LABEL, BUTTON_HEIGHT_LARGE))
        edit_btn.clicked.connect(self.edit_template)
        buttons_layout.addWidget(edit_btn)
        
        delete_btn = QPushButton("🗑️ حذف")
        delete_btn.setMinimumHeight(BUTTON_HEIGHT_LARGE)
        delete_btn.setStyleSheet(get_button_style(COLOR_DANGER, FONT_SIZE_LABEL, BUTTON_HEIGHT_LARGE))
        delete_btn.clicked.connect(self.delete_template)
        buttons_layout.addWidget(delete_btn)
        
        layout.addLayout(buttons_layout)
    
    def load_templates(self):
        """بارگذاری Template ها"""
        try:
            self.templates_list.clear()
            
            templates = db_manager.get_all_templates()
            
            for template in templates:
                mappings = template.column_mappings or {}
                formulas = sum(1 for m in mappings.values() if isinstance(m, dict) and m.get('formula'))
                
                item_text = f"""
📋 {template.name}
   📄 File: {Path(template.template_path).name if template.template_path else 'N/A'}
   🔗 Mappings: {len(mappings)} | ⚡ Formulas: {formulas}
   🏷️ Type: {template.template_type}
                """.strip()
                
                self.templates_list.addItem(item_text)
                self.templates_list.item(self.templates_list.count() - 1).setData(Qt.ItemDataRole.UserRole, template)
        
        except Exception as e:
            app_logger.error(f"Error loading templates: {e}")
    
    def add_template(self):
        """افزودن Template جدید"""
        wizard = TemplateWizard(self)
        if wizard.exec():
            self.load_templates()
    
    def edit_template(self):
        """ویرایش Template"""
        current_item = self.templates_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "هشدار", "لطفاً یک Template انتخاب کنید")
            return
        
        template = current_item.data(Qt.ItemDataRole.UserRole)
        wizard = TemplateWizard(self, template=template)
        if wizard.exec():
            self.load_templates()
    
    def delete_template(self):
        """حذف Template"""
        current_item = self.templates_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "هشدار", "لطفاً یک Template انتخاب کنید")
            return
        
        template = current_item.data(Qt.ItemDataRole.UserRole)
        
        reply = QMessageBox.question(
            self,
            "تأیید حذف",
            f"آیا از حذف Template '{template.name}' مطمئن هستید؟",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            success, message = db_manager.delete_template(template.id)
            if success:
                QMessageBox.information(self, "موفقیت", "Template با موفقیت حذف شد")
                self.load_templates()
            else:
                QMessageBox.critical(self, "خطا", f"خطا در حذف Template:\n{message}")
