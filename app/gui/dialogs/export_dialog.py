"""
دیالوگ تولید خروجی Excel
"""
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLineEdit, QPushButton, QCheckBox, QLabel,
    QMessageBox, QGroupBox, QRadioButton, QButtonGroup,
    QProgressDialog, QFileDialog
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont
from pathlib import Path
from datetime import datetime

from app.core.database import db_manager
from app.core.logger import app_logger


class ExportThread(QThread):
    """Thread برای تولید خروجی"""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(bool, str, str)
    
    def __init__(self, export_type, output_path):
        super().__init__()
        self.export_type = export_type
        self.output_path = output_path
        self.logger = app_logger
    
    def run(self):
        """اجرای تولید خروجی"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            self.progress.emit(10, "در حال دریافت داده‌ها...")
            
            # دریافت داده‌های استخراج نشده
            data_list = db_manager.get_unexported_data(self.export_type)
            
            if not data_list:
                self.finished.emit(False, "هیچ داده‌ای برای خروجی یافت نشد!", "")
                return
            
            self.progress.emit(30, f"تعداد {len(data_list)} رکورد یافت شد...")
            
            # تبدیل به DataFrame
            records = []
            for item in data_list:
                record = item.data.copy()
                record['_id'] = item.id
                record['_created_at'] = item.created_at.strftime("%Y-%m-%d %H:%M:%S")
                records.append(record)
            
            df = pd.DataFrame(records)
            
            self.progress.emit(50, "در حال ایجاد فایل Excel...")
            
            # ایجاد Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # استایل هدر
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # نوشتن هدر
            for col_idx, column in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            self.progress.emit(70, "در حال نوشتن داده‌ها...")
            
            # نوشتن داده‌ها
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="right")
            
            # تنظیم عرض ستون‌ها
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # فریز کردن سرتیتر
            ws.freeze_panes = "A2"
            
            # افزودن فیلتر
            ws.auto_filter.ref = ws.dimensions
            
            self.progress.emit(90, "در حال ذخیره فایل...")
            
            # ذخیره فایل
            wb.save(self.output_path)
            
            # علامت‌گذاری رکوردها
            data_ids = [item.id for item in data_list]
            db_manager.mark_as_exported(data_ids, self.export_type)
            
            self.progress.emit(100, "تمام!")
            self.finished.emit(True, f"✅ {len(data_list)} رکورد با موفقیت خروجی گرفته شد!", self.output_path)
            
        except Exception as e:
            self.logger.error(f"خطا در تولید خروجی: {str(e)}")
            self.finished.emit(False, f"❌ خطا: {str(e)}", "")


class ExportDialog(QDialog):
    """دیالوگ تولید خروجی"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.logger = app_logger
        self.export_thread = None
        self.init_ui()
    
    def init_ui(self):
        """راه‌اندازی رابط کاربری"""
        self.setWindowTitle("📤 تولید خروجی Excel")
        self.setMinimumWidth(500)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # عنوان
        title_label = QLabel("📤 تولید فایل Excel")
        title_label.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        title_label.setStyleSheet("color: #2196F3; padding: 10px;")
        layout.addWidget(title_label)
        
        # انتخاب نوع خروجی
        type_group = QGroupBox("📋 نوع خروجی")
        type_layout = QVBoxLayout()
        
        self.type_group = QButtonGroup()
        
        type1_radio = QRadioButton("📄 نوع 1 - فروش عمومی")
        type1_radio.setChecked(True)
        self.type_group.addButton(type1_radio, 1)
        type_layout.addWidget(type1_radio)
        
        type2_radio = QRadioButton("📄 نوع 2 - گزارش مالی")
        self.type_group.addButton(type2_radio, 2)
        type_layout.addWidget(type2_radio)
        
        type3_radio = QRadioButton("📄 نوع 3 - گزارش جامع")
        self.type_group.addButton(type3_radio, 3)
        type_layout.addWidget(type3_radio)
        
        type_group.setLayout(type_layout)
        layout.addWidget(type_group)
        
        # مسیر خروجی
        path_group = QGroupBox("📁 مسیر ذخیره")
        path_layout = QHBoxLayout()
        
        self.output_path_input = QLineEdit()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_path = f"data/exports/export_{timestamp}.xlsx"
        self.output_path_input.setText(default_path)
        path_layout.addWidget(self.output_path_input)
        
        browse_btn = QPushButton("📁 انتخاب")
        browse_btn.clicked.connect(self.browse_output_path)
        path_layout.addWidget(browse_btn)
        
        path_group.setLayout(path_layout)
        layout.addWidget(path_group)
        
        # آمار
        self.stats_label = QLabel("در حال بارگذاری آمار...")
        self.stats_label.setStyleSheet("""
            background: #e3f2fd;
            padding: 15px;
            border-radius: 5px;
            border: 2px solid #2196F3;
            font-weight: bold;
        """)
        layout.addWidget(self.stats_label)
        
        # بارگذاری آمار
        self.load_stats()
        
        # دکمه‌ها
        buttons_layout = QHBoxLayout()
        
        export_btn = QPushButton("📤 تولید خروجی")
        export_btn.setStyleSheet(self.get_button_style("#4CAF50"))
        export_btn.clicked.connect(self.start_export)
        buttons_layout.addWidget(export_btn)
        
        cancel_btn = QPushButton("❌ انصراف")
        cancel_btn.setStyleSheet(self.get_button_style("#F44336"))
        cancel_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(cancel_btn)
        
        layout.addLayout(buttons_layout)
    
    def load_stats(self):
        """بارگذاری آمار"""
        try:
            stats = db_manager.get_statistics()
            
            total = stats.get('total_records', 0)
            exported = stats.get('exported_records', 0)
            pending = stats.get('pending_records', 0)
            
            self.stats_label.setText(
                f"📊 آمار:\n"
                f"• کل رکوردها: {total:,}\n"
                f"• خروجی گرفته شده: {exported:,}\n"
                f"• در انتظار: {pending:,}"
            )
        except Exception as e:
            self.stats_label.setText(f"❌ خطا در بارگذاری آمار: {str(e)}")
    
    def browse_output_path(self):
        """انتخاب مسیر خروجی"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "انتخاب مسیر ذخیره",
            self.output_path_input.text(),
            "Excel Files (*.xlsx)"
        )
        
        if file_path:
            self.output_path_input.setText(file_path)
    
    def start_export(self):
        """شروع تولید خروجی"""
        output_path = self.output_path_input.text()
        
        if not output_path:
            QMessageBox.warning(self, "خطا", "لطفاً مسیر خروجی را مشخص کنید!")
            return
        
        # ایجاد دایرکتوری
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        # نوع خروجی
        export_type = self.type_group.checkedId()
        
        # نمایش پیشرفت
        progress = QProgressDialog("در حال تولید خروجی...", "لغو", 0, 100, self)
        progress.setWindowTitle("تولید خروجی")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        
        # ایجاد thread
        self.export_thread = ExportThread(export_type, output_path)
        
        # اتصال سیگنال‌ها
        self.export_thread.progress.connect(
            lambda value, msg: (progress.setValue(value), progress.setLabelText(msg))
        )
        self.export_thread.finished.connect(
            lambda success, msg, path: self.on_export_finished(success, msg, path, progress)
        )
        
        # شروع
        self.export_thread.start()
    
    def on_export_finished(self, success, message, output_path, progress_dialog):
        """پایان تولید خروجی"""
        progress_dialog.close()
        
        if success:
            reply = QMessageBox.information(
                self,
                "موفق",
                f"{message}\n\nآیا می‌خواهید فایل را باز کنید؟",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                import os
                os.startfile(output_path)
            
            self.accept()
        else:
            QMessageBox.critical(self, "خطا", message)
    
    def get_button_style(self, color):
        """استایل دکمه"""
        return f"""
            QPushButton {{
                background: {color};
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 11pt;
                font-weight: bold;
                min-width: 100px;
            }}
            QPushButton:hover {{
                opacity: 0.9;
            }}
        """
