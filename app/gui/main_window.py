"""
پنجره اصلی برنامه GT-Land Manager
"""
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QLabel, QTabWidget, QMessageBox,
    QStatusBar, QMenuBar, QMenu, QToolBar, QFrame,
    QDialog
)
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QAction, QIcon, QFont

from app.core.database import db_manager
from app.core.logger import app_logger
from app.gui.dialogs.sheet_config_dialog import SheetConfigDialog
from app.gui.dialogs.advanced_export_dialog import AdvancedExportDialog
from app.gui.dialogs.settings_dialog import SettingsDialog
from app.gui.dialogs.template_manager_dialog_advanced import TemplateManagerDialog
from app.gui.widgets.sheet_list_widget import SheetListWidget
from app.gui.widgets.extraction_widget import ExtractionWidget
from app.gui.widgets.reports_widget import ReportsWidget
from app.gui.widgets.data_viewer_widget import DataViewerWidget
from app.gui.financial.smart_import_wizard import SmartImportWizard
from app.gui.financial.comprehensive_reports_widget import ComprehensiveReportsWidget
from app.gui.financial.inventory_management_widget import InventoryManagementWidget
from app.gui.financial.per_sheet_mapping_dialog import PerSheetFieldMappingDialog
from app.utils.ui_constants import (
    FONT_SIZE_TITLE, FONT_SIZE_SECTION, FONT_SIZE_BUTTON,
    BUTTON_HEIGHT_LARGE, BUTTON_HEIGHT_MEDIUM,
    SPACING_LARGE, SPACING_MEDIUM, MARGIN_LARGE,
    COLOR_PRIMARY, COLOR_SUCCESS, COLOR_WARNING, COLOR_DANGER,
    get_button_style
)


class MainWindow(QMainWindow):
    """پنجره اصلی برنامه"""
    
    def __init__(self):
        super().__init__()
        self.logger = app_logger
        self.init_ui()
        self.load_statistics()
    
    def init_ui(self):
        """راه‌اندازی رابط کاربری"""
        # تنظیمات پنجره
        self.setWindowTitle("GT-Land Manager - مدیریت داده‌های فروش")
        self.setGeometry(100, 100, 1400, 800)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        # ایجاد منو
        self.create_menu()
        
        # ایجاد نوار ابزار
        self.create_toolbar()
        
        # ایجاد ویجت مرکزی
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # لی‌اوت اصلی
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # عنوان
        title_label = QLabel("🚀 GT-Land Manager")
        title_font = QFont("Segoe UI", FONT_SIZE_TITLE, QFont.Weight.Bold)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet(f"color: {COLOR_PRIMARY}; padding: 15px;")
        main_layout.addWidget(title_label)
        
        # تب‌ها
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #e0e0e0;
                border-radius: 5px;
                background: white;
            }
            QTabBar::tab {
                background: #f5f5f5;
                border: 1px solid #e0e0e0;
                padding: 10px 20px;
                margin: 2px;
                border-radius: 5px;
                font-size: 12pt;
            }
            QTabBar::tab:selected {
                background: #2196F3;
                color: white;
            }
            QTabBar::tab:hover {
                background: #64B5F6;
            }
        """)
        
        # تب داشبورد
        self.create_dashboard_tab()
        
        # تب مدیریت شیت‌ها
        self.create_sheets_tab()
        
        # تب استخراج داده
        self.create_extract_tab()
        
        # تب تولید خروجی
        self.create_export_tab()
        
        # تب گزارش‌ها
        self.create_reports_tab()
        # تب تنظیمات
        self.create_settings_tab()
        
        # تب گزارش‌ساز هوشمند
        self.create_report_builder_tab()
        
        # تب گزارشات جامع
        self.create_comprehensive_reports_tab()
        
        # تب مدیریت موجودی
        self.create_inventory_management_tab()
        
        # تب مدیریت BI (Smart Import + Field Mapping)
        self.create_bi_platform_tab()
        
        main_layout.addWidget(self.tabs)
        
        # نوار وضعیت
        self.statusBar().showMessage("✅ آماده")
        self.statusBar().setStyleSheet("""
            QStatusBar {
                background: #f5f5f5;
                color: #333;
                font-size: 10pt;
                padding: 5px;
            }
        """)
    
    def create_menu(self):
        """ایجاد منوی اصلی"""
        menubar = self.menuBar()
        
        # منوی فایل
        file_menu = menubar.addMenu("📁 فایل")
        
        exit_action = QAction("🚪 خروج", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # منوی ابزارها
        tools_menu = menubar.addMenu("🔧 ابزارها")
        
        refresh_action = QAction("🔄 بروزرسانی", self)
        refresh_action.setShortcut("F5")
        refresh_action.triggered.connect(self.refresh_data)
        tools_menu.addAction(refresh_action)
        
        # مدیریت نقش‌های فیلدها (مخصوص هر شیت)
        per_sheet_mapping_action = QAction("🗂️ مدیریت نقش‌های فیلدها", self)
        per_sheet_mapping_action.triggered.connect(self.open_per_sheet_mapping)
        tools_menu.addAction(per_sheet_mapping_action)
        
        # افزودن ویزارد import داده
        import_wizard_action = QAction("🔄 ورود داده از شیت‌ها", self)
        import_wizard_action.triggered.connect(self.open_smart_import_wizard)
        tools_menu.addAction(import_wizard_action)
        
        # منوی راهنما
        help_menu = menubar.addMenu("❓ راهنما")
        
        about_action = QAction("ℹ️ درباره", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
    
    def create_toolbar(self):
        """ایجاد نوار ابزار"""
        toolbar = QToolBar()
        toolbar.setIconSize(QSize(32, 32))
        toolbar.setMovable(False)
        toolbar.setStyleSheet("""
            QToolBar {
                background: #fafafa;
                border-bottom: 2px solid #e0e0e0;
                spacing: 10px;
                padding: 5px;
            }
            QToolButton {
                background: white;
                border: 1px solid #e0e0e0;
                border-radius: 5px;
                padding: 8px;
                margin: 2px;
            }
            QToolButton:hover {
                background: #e3f2fd;
                border-color: #2196F3;
            }
        """)
        
        # دکمه‌های سیستم جدید
        per_sheet_mapping_btn = QPushButton("🗂️ نقش‌های فیلدها")
        per_sheet_mapping_btn.clicked.connect(self.open_per_sheet_mapping)
        per_sheet_mapping_btn.setToolTip("مدیریت نقش‌های فیلدها به تفکیک شیت")
        toolbar.addWidget(per_sheet_mapping_btn)
        
        smart_import_btn = QPushButton("🚀 Import هوشمند")
        smart_import_btn.clicked.connect(self.open_smart_import)
        smart_import_btn.setToolTip("Import با گروه‌بندی خودکار")
        toolbar.addWidget(smart_import_btn)
        
        # conflicts_btn = QPushButton("⚠️ تداخل‌ها")
        # conflicts_btn.clicked.connect(self.open_conflicts)
        # conflicts_btn.setToolTip("مدیریت تداخل‌ها")
        # toolbar.addWidget(conflicts_btn)
        # ⚠️ غیرفعال شد - با RawData قدیمی کار می‌کند
        
        toolbar.addSeparator()
        
        # دکمه بروزرسانی
        refresh_btn = QPushButton("🔄 بروزرسانی")
        refresh_btn.clicked.connect(self.refresh_data)
        toolbar.addWidget(refresh_btn)
        
        toolbar.addSeparator()
        
        # دکمه خروج
        exit_btn = QPushButton("🚪 خروج")
        exit_btn.clicked.connect(self.close)
        toolbar.addWidget(exit_btn)
        
        self.addToolBar(toolbar)
    
    def create_dashboard_tab(self):
        """ایجاد تب داشبورد"""
        dashboard = QWidget()
        layout = QVBoxLayout(dashboard)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # عنوان
        title = QLabel("📊 داشبورد اطلاعات")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #333;")
        layout.addWidget(title)
        
        # کارت‌های آمار
        stats_layout = QHBoxLayout()
        
        # تعداد تنظیمات
        self.configs_card = self.create_stat_card("📝 تنظیمات شیت‌ها", "0", "#2196F3")
        stats_layout.addWidget(self.configs_card)
        
        # تعداد رکوردها
        self.records_card = self.create_stat_card("📦 کل رکوردها", "0", "#4CAF50")
        stats_layout.addWidget(self.records_card)
        
        # خروجی گرفته شده
        self.exported_card = self.create_stat_card("✅ خروجی گرفته شده", "0", "#FF9800")
        stats_layout.addWidget(self.exported_card)
        
        # در انتظار
        self.pending_card = self.create_stat_card("⏳ در انتظار", "0", "#F44336")
        stats_layout.addWidget(self.pending_card)
        
        layout.addLayout(stats_layout)
        
        # دکمه‌های عملیات سریع
        quick_actions = QHBoxLayout()
        quick_actions.setSpacing(SPACING_MEDIUM)
        
        extract_btn = QPushButton("📥 استخراج سریع")
        extract_btn.setMinimumHeight(BUTTON_HEIGHT_LARGE)
        extract_btn.setStyleSheet(get_button_style(COLOR_PRIMARY, FONT_SIZE_BUTTON, BUTTON_HEIGHT_LARGE))
        extract_btn.clicked.connect(lambda: self.tabs.setCurrentIndex(2))
        quick_actions.addWidget(extract_btn)
        
        export_btn = QPushButton("📤 تولید خروجی")
        export_btn.setMinimumHeight(BUTTON_HEIGHT_LARGE)
        export_btn.setStyleSheet(get_button_style(COLOR_SUCCESS, FONT_SIZE_BUTTON, BUTTON_HEIGHT_LARGE))
        export_btn.clicked.connect(lambda: self.tabs.setCurrentIndex(3))
        quick_actions.addWidget(export_btn)
        
        layout.addLayout(quick_actions)
        
        layout.addStretch()
        
        self.tabs.addTab(dashboard, "📊 داشبورد")
    
    def create_sheets_tab(self):
        """ایجاد تب مدیریت شیت‌ها"""
        sheets = QWidget()
        layout = QVBoxLayout(sheets)
        
        title = QLabel("📋 مدیریت تنظیمات Google Sheets")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        layout.addWidget(title)
        
        # افزودن ویجت لیست
        self.sheet_list_widget = SheetListWidget()
        layout.addWidget(self.sheet_list_widget)
        
        self.tabs.addTab(sheets, "📋 مدیریت شیت‌ها")
    
    def create_extract_tab(self):
        """ایجاد تب استخراج داده"""
        # استفاده از ویجت استخراج
        self.extraction_widget = ExtractionWidget()
        self.tabs.addTab(self.extraction_widget, "📥 استخراج داده")
    
    def create_export_tab(self):
        """ایجاد تب مدیریت داده‌ها و انتقال به مرحله بعد"""
        export = QWidget()
        layout = QVBoxLayout(export)
        
        title = QLabel("📊 مدیریت داده‌ها و انتقال")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        layout.addWidget(title)
        
        # ویجت نمایش داده‌ها (شامل دکمه انتقال به Stage 2)
        self.data_viewer_widget = DataViewerWidget()
        layout.addWidget(self.data_viewer_widget)
        
        self.tabs.addTab(export, "📊 مدیریت داده‌ها")
    
    def create_reports_tab(self):
        """ایجاد تب گزارش‌ها"""
        # استفاده از ویجت گزارش‌ها
        self.reports_widget = ReportsWidget()
        self.tabs.addTab(self.reports_widget, "📈 گزارش‌ها")
    
    def create_settings_tab(self):
        """ایجاد تب تنظیمات"""
        settings = QWidget()
        layout = QVBoxLayout(settings)
        
        title = QLabel("⚙️ تنظیمات برنامه")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        layout.addWidget(title)
        
        # دکمه باز کردن تنظیمات
        settings_btn = QPushButton("⚙️ باز کردن تنظیمات")
        settings_btn.setMinimumHeight(BUTTON_HEIGHT_LARGE)
        settings_btn.setStyleSheet(get_button_style(COLOR_PRIMARY, FONT_SIZE_SECTION, BUTTON_HEIGHT_LARGE))
        settings_btn.clicked.connect(self.open_settings)
        layout.addWidget(settings_btn)
        
        # اطلاعات
        info_label = QLabel(
            "🔧 تنظیمات عمومی\n"
            "🔐 تنظیمات Google Sheets\n"
            "💾 تنظیمات دیتابیس\n"
            "📤 تنظیمات خروجی"
        )
        info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_label.setStyleSheet("padding: 50px; font-size: 14pt; color: #666; line-height: 2;")
        layout.addWidget(info_label)
        
        layout.addStretch()
        
        self.tabs.addTab(settings, "⚙️ تنظیمات")
    
    def create_report_builder_tab(self):
        """ایجاد تب گزارش‌ساز هوشمند"""
        from app.gui.financial.report_builder_widget import ReportBuilderWidget
        
        self.report_builder_widget = ReportBuilderWidget()
        self.report_builder_widget.export_requested.connect(self.handle_report_export)
        
        self.tabs.addTab(self.report_builder_widget, "📊 گزارش‌ساز هوشمند")
    
    def create_comprehensive_reports_tab(self):
        """ایجاد تب گزارشات جامع (10 نوع گزارش)"""
        self.comprehensive_reports_widget = ComprehensiveReportsWidget()
        self.tabs.addTab(self.comprehensive_reports_widget, "📊 گزارشات جامع")
    
    def create_inventory_management_tab(self):
        """ایجاد تب مدیریت موجودی (Inventory Management)"""
        self.inventory_management_widget = InventoryManagementWidget()
        self.tabs.addTab(self.inventory_management_widget, "📦 مدیریت موجودی")
    
    def create_bi_platform_tab(self):
        """ایجاد تب مدیریت BI (Smart Import + Tools)"""
        from PyQt6.QtWidgets import QWidget, QVBoxLayout, QPushButton, QGroupBox, QLabel
        
        bi_widget = QWidget()
        layout = QVBoxLayout(bi_widget)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # عنوان
        title = QLabel("🔄 مدیریت BI - ابزارهای Import و تنظیمات")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #2196F3; padding: 15px;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        # گروه Smart Import
        import_group = QGroupBox("📥 Import داده‌ها")
        import_group.setStyleSheet("""
            QGroupBox {
                font-size: 12pt;
                font-weight: bold;
                border: 2px solid #2196F3;
                border-radius: 10px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 20px;
                padding: 0 10px;
            }
        """)
        import_layout = QVBoxLayout()
        import_layout.setSpacing(10)
        
        # دکمه Smart Import
        smart_import_btn = QPushButton("🚀 Smart Import Wizard")
        smart_import_btn.setMinimumHeight(50)
        smart_import_btn.setStyleSheet("""
            QPushButton {
                background: #4CAF50;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 13pt;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover {
                background: #45a049;
            }
            QPushButton:pressed {
                background: #3d8b40;
            }
        """)
        smart_import_btn.clicked.connect(self.open_smart_import_wizard)
        import_layout.addWidget(smart_import_btn)
        
        # توضیحات
        desc = QLabel("📝 Import هوشمند با تشخیص خودکار ستون‌ها و مدیریت تداخل")
        desc.setStyleSheet("color: #666; font-size: 10pt; padding: 5px;")
        desc.setWordWrap(True)
        import_layout.addWidget(desc)
        
        import_group.setLayout(import_layout)
        layout.addWidget(import_group)
        
        # گروه تنظیمات Field Mapping
        mapping_group = QGroupBox("🗂️ تنظیمات نقش‌ها")
        mapping_group.setStyleSheet("""
            QGroupBox {
                font-size: 12pt;
                font-weight: bold;
                border: 2px solid #FF9800;
                border-radius: 10px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 20px;
                padding: 0 10px;
            }
        """)
        mapping_layout = QVBoxLayout()
        mapping_layout.setSpacing(10)
        
        # دکمه Per-Sheet Mapping
        field_mapping_btn = QPushButton("🗂️ مدیریت نقش‌های فیلدها")
        field_mapping_btn.setMinimumHeight(50)
        field_mapping_btn.setStyleSheet("""
            QPushButton {
                background: #FF9800;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 13pt;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover {
                background: #FB8C00;
            }
            QPushButton:pressed {
                background: #F57C00;
            }
        """)
        field_mapping_btn.clicked.connect(self.open_per_sheet_mapping)
        mapping_layout.addWidget(field_mapping_btn)
        
        # توضیحات
        desc2 = QLabel("⚙️ تنظیم نقش‌های فیلدها برای هر شیت (Purchase, Sale, Bonus)")
        desc2.setStyleSheet("color: #666; font-size: 10pt; padding: 5px;")
        desc2.setWordWrap(True)
        mapping_layout.addWidget(desc2)
        
        mapping_group.setLayout(mapping_layout)
        layout.addWidget(mapping_group)
        
        # گروه گزارش حرفه‌ای (Professional Grid)
        grid_group = QGroupBox("📊 گزارش حرفه‌ای")
        grid_group.setStyleSheet("""
            QGroupBox {
                font-size: 12pt;
                font-weight: bold;
                border: 2px solid #9C27B0;
                border-radius: 10px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 20px;
                padding: 0 10px;
            }
        """)
        grid_layout = QVBoxLayout()
        grid_layout.setSpacing(10)
        
        # دکمه Professional Grid
        grid_btn = QPushButton("📊 نمایش گزارش Excel-like")
        grid_btn.setMinimumHeight(50)
        grid_btn.setStyleSheet("""
            QPushButton {
                background: #9C27B0;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 13pt;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover {
                background: #8E24AA;
            }
            QPushButton:pressed {
                background: #7B1FA2;
            }
        """)
        grid_btn.clicked.connect(self.open_professional_grid)
        grid_layout.addWidget(grid_btn)
        
        # توضیحات
        desc3 = QLabel("💎 نمایش داده‌ها به صورت حرفه‌ای: خرید، فروش به تفکیک Platform، سود/زیان")
        desc3.setStyleSheet("color: #666; font-size: 10pt; padding: 5px;")
        desc3.setWordWrap(True)
        grid_layout.addWidget(desc3)
        
        grid_group.setLayout(grid_layout)
        layout.addWidget(grid_group)
        
        # گروه مدیریت تداخل - غیرفعال شد
        # conflict_group = QGroupBox("⚠️ مدیریت تداخل")
        # ... با RawData قدیمی کار می‌کند - باید بازنویسی شود
        
        # فضای خالی
        layout.addStretch()
        
        # راهنما
        help_label = QLabel(
            "💡 راهنما:\n"
            "1. ابتدا نقش‌های فیلدها را برای هر شیت تنظیم کنید\n"
            "2. سپس از Smart Import Wizard برای Import استفاده کنید\n"
            "3. در صورت بروز تداخل، از بخش مدیریت تداخل استفاده کنید"
        )
        help_label.setStyleSheet("""
            background: #E3F2FD;
            border: 2px solid #2196F3;
            border-radius: 8px;
            padding: 15px;
            color: #1565C0;
            font-size: 10pt;
        """)
        help_label.setWordWrap(True)
        layout.addWidget(help_label)
        
        self.tabs.addTab(bi_widget, "🔄 مدیریت BI")
    
    def handle_report_export(self, export_data):
        """مدیریت Export گزارش"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # ایجاد فایل Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "گزارش"
            
            # عنوان
            ws['A1'] = export_data.get('report_type', 'گزارش')
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            # تاریخ
            ws['A2'] = f"تاریخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            
            # داده‌ها (فعلاً ساده)
            ws['A4'] = "این گزارش بعد از تکمیل Export به Excel صادر می‌شود"
            
            # ذخیره
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            QMessageBox.information(
                self,
                "Export موفق",
                f"✅ گزارش ذخیره شد:\n{filename}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در Export:\n{e}")
    

    def create_stat_card(self, title, value, color):
        """ایجاد کارت آمار"""
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background: white;
                border: 2px solid {color};
                border-radius: 10px;
                padding: 20px;
            }}
        """)
        
        layout = QVBoxLayout(frame)
        
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 12))
        title_label.setStyleSheet(f"color: {color}; font-weight: bold;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        value_label = QLabel(value)
        value_label.setFont(QFont("Segoe UI", 32, QFont.Weight.Bold))
        value_label.setStyleSheet(f"color: {color};")
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        value_label.setObjectName("value_label")
        layout.addWidget(value_label)
        
        return frame
    
    def load_statistics(self):
        """بارگذاری آمار"""
        try:
            stats = db_manager.get_statistics()
            
            # بروزرسانی کارت‌ها
            self.update_card_value(self.configs_card, str(stats.get('total_configs', 0)))
            self.update_card_value(self.records_card, str(stats.get('total_records', 0)))
            self.update_card_value(self.exported_card, str(stats.get('exported_records', 0)))
            self.update_card_value(self.pending_card, str(stats.get('pending_records', 0)))
            
            self.logger.success("آمار بروزرسانی شد")
        except Exception as e:
            self.logger.error(f"خطا در بارگذاری آمار: {str(e)}")
    
    def update_card_value(self, card, value):
        """بروزرسانی مقدار کارت"""
        value_label = card.findChild(QLabel, "value_label")
        if value_label:
            value_label.setText(value)
    
    def refresh_data(self):
        """بروزرسانی داده‌ها"""
        self.load_statistics()
        self.statusBar().showMessage("✅ داده‌ها بروزرسانی شدند", 3000)
        QMessageBox.information(self, "بروزرسانی", "✅ داده‌ها با موفقیت بروزرسانی شدند!")
    
    def add_sheet_config(self):
        """افزودن تنظیمات شیت جدید"""
        dialog = SheetConfigDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_data()
    
    def open_export_dialog(self):
        """باز کردن دیالوگ تولید خروجی"""
        dialog = AdvancedExportDialog(self)
        dialog.export_completed.connect(self.on_export_completed)
        dialog.exec()
    
    def open_template_manager(self):
        """باز کردن مدیریت Template ها"""
        dialog = TemplateManagerDialog(self)
        dialog.exec()
    
    def on_export_completed(self, success, message):
        """بعد از Export"""
        if success:
            # بروزرسانی Data Viewer
            if hasattr(self, 'data_viewer_widget'):
                self.data_viewer_widget.refresh_data()
            
            # بروزرسانی آمار
            self.load_statistics()
        
        self.refresh_data()
    
    def open_settings(self):
        """باز کردن تنظیمات"""
        dialog = SettingsDialog(self)
        dialog.exec()
    
    def show_about(self):
        """نمایش درباره برنامه"""
        about_text = """
        <h2>GT-Land Manager</h2>
        <p><b>نسخه:</b> Ver 9</p>
        <p><b>توضیحات:</b> سیستم مدیریت و همگام‌سازی داده‌های فروش</p>
        <br>
        <p>✅ استخراج هوشمند از Google Sheets</p>
        <p>✅ ذخیره در دیتابیس SQLite</p>
        <p>✅ تولید خروجی Excel حرفه‌ای</p>
        <p>✅ رابط کاربری گرافیکی کامل</p>
        <p>✅ سیستم آرشیو و تمیزسازی</p>
        <br>
        <p><b>👨‍💻 توسعه‌دهنده:</b> علیرضا حامد</p>
        <p><b>📅 سال:</b> 2025</p>
        <p><b>📧 پشتیبانی:</b> GT-Land Team</p>
        """
        QMessageBox.about(self, "درباره برنامه", about_text)
    
    def open_smart_import_wizard(self):
        """باز کردن ویزارد ورود هوشمند داده از Google Sheets"""
        try:
            from app.gui.financial.smart_import_wizard import SmartImportWizard
            
            dialog = SmartImportWizard(self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                # بروزرسانی لیست شیت‌ها در تب Import
                if hasattr(self, 'data_viewer_widget'):
                    self.data_viewer_widget.load_sheets()
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "خطا", f"❌ خطا در باز کردن ویزارد:\n{str(e)}")
    
    def open_professional_grid(self):
        """باز کردن گزارش حرفه‌ای Excel-like"""
        try:
            from app.gui.financial.professional_grid_widget import ProfessionalGridWidget
            
            # ساخت پنجره جداگانه
            grid_window = QDialog(self)
            grid_window.setWindowTitle("📊 گزارش حرفه‌ای - خرید و فروش")
            grid_window.resize(1400, 800)
            
            layout = QVBoxLayout(grid_window)
            layout.setContentsMargins(0, 0, 0, 0)
            
            # افزودن ویجت Grid
            grid_widget = ProfessionalGridWidget()
            layout.addWidget(grid_widget)
            
            grid_window.exec()
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "خطا", f"❌ خطا در باز کردن گزارش:\n{str(e)}")
    

    def open_per_sheet_mapping(self):
        """باز کردن مدیریت نقش‌ها به تفکیک شیت - سیستم جدیدتر!"""
        try:
            dialog = PerSheetFieldMappingDialog(self)
            dialog.exec()
            self.statusBar().showMessage("✅ نقش‌های شیت بروزرسانی شد", 3000)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "خطا", f"❌ خطا در باز کردن مدیر نقش‌های شیت:\n{str(e)}")
    
    def open_smart_import(self):
        """باز کردن Import هوشمند - سیستم جدید"""
        try:
            dialog = SmartImportWizard(self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self.statusBar().showMessage("✅ Import با موفقیت کامل شد", 5000)
                self.refresh_data()
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "خطا", f"❌ خطا در Import:\n{str(e)}")
    
    def open_conflicts(self):
        """باز کردن مدیریت تداخل‌ها - DEPRECATED"""
        QMessageBox.information(
            self, "⚠️ غیرفعال",
            "این بخش با سیستم قدیمی کار می‌کند و غیرفعال شده است.\n\n"
            "از Smart Import Wizard برای Import استفاده کنید."
        )
    
    def closeEvent(self, event):
        """رویداد بستن برنامه"""
        reply = QMessageBox.question(
            self,
            "خروج",
            "آیا می‌خواهید از برنامه خارج شوید؟",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.logger.info("برنامه بسته شد")
            event.accept()
        else:
            event.ignore()
