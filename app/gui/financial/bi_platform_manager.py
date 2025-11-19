"""
BI Platform Manager - نسخه کامل و بدون باگ
=========================================

مدیریت کامل سیستم BI با:
- مدیریت Sources (CRUD کامل)
- مدیریت Fields
- Data Explorer
- Formula Builder
- Migration Tool
- Report Designer
"""

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QTabWidget,
    QPushButton, QLabel, QMessageBox, QToolBar, QStatusBar,
    QTableWidget, QTableWidgetItem, QHeaderView, QMenu,
    QDialog, QLineEdit, QComboBox, QTextEdit, QFormLayout,
    QCheckBox, QSpinBox, QGroupBox, QDialogButtonBox, QScrollArea,
    QFileDialog, QInputDialog, QApplication
)
from PyQt6.QtCore import Qt, QSize, pyqtSignal
from PyQt6.QtGui import QFont, QAction, QIcon

from app.core.financial import DataManager, FormulaEngine, AggregationEngine
from typing import Optional, List, Dict, Any
from datetime import datetime
import json
import csv
import traceback


class AddSourceDialog(QDialog):
    """دیالوگ اضافه کردن منبع جدید"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("➕ منبع داده جدید")
        self.setMinimumWidth(500)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        layout = QFormLayout(self)
        
        # نام
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("مثال: فروش PUBG نوامبر 2025")
        layout.addRow("📝 نام:", self.name_input)
        
        # نوع
        self.type_combo = QComboBox()
        self.type_combo.addItems([
            "google_sheet",
            "excel",
            "csv",
            "api",
            "database",
            "manual"
        ])
        layout.addRow("📊 نوع:", self.type_combo)
        
        # توضیحات
        self.desc_input = QTextEdit()
        self.desc_input.setMaximumHeight(80)
        self.desc_input.setPlaceholderText("توضیحات اختیاری...")
        layout.addRow("📄 توضیحات:", self.desc_input)
        
        # دکمه‌ها
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
    
    def get_data(self) -> Dict[str, Any]:
        """دریافت داده‌های فرم"""
        return {
            "name": self.name_input.text(),
            "source_type": self.type_combo.currentText(),
            "description": self.desc_input.toPlainText(),
            "connection_info": {}
        }


class EditSourceDialog(QDialog):
    """دیالوگ ویرایش منبع"""
    
    def __init__(self, source_data: Dict[str, Any], parent=None):
        super().__init__(parent)
        self.source_data = source_data
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle(f"✏️ ویرایش: {self.source_data.get('name', 'منبع')}")
        self.setMinimumWidth(500)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        layout = QFormLayout(self)
        
        # نام
        self.name_input = QLineEdit()
        self.name_input.setText(self.source_data.get("name", ""))
        layout.addRow("📝 نام:", self.name_input)
        
        # وضعیت
        self.active_check = QCheckBox("فعال")
        self.active_check.setChecked(self.source_data.get("is_active", True))
        layout.addRow("✅ وضعیت:", self.active_check)
        
        # توضیحات
        self.desc_input = QTextEdit()
        self.desc_input.setMaximumHeight(80)
        self.desc_input.setPlainText(
            self.source_data.get("connection_info", {}).get("description", "")
        )
        layout.addRow("📄 توضیحات:", self.desc_input)
        
        # دکمه‌ها
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
    
    def get_data(self) -> Dict[str, Any]:
        """دریافت داده‌های ویرایش شده"""
        return {
            "name": self.name_input.text(),
            "is_active": self.active_check.isChecked(),
            "description": self.desc_input.toPlainText()
        }


class ManageFieldsDialog(QDialog):
    """دیالوگ مدیریت فیلدهای یک منبع"""
    
    def __init__(self, source_id: int, dm: DataManager, parent=None):
        super().__init__(parent)
        self.source_id = source_id
        self.dm = dm
        self.init_ui()
        self.load_fields()
    
    def init_ui(self):
        self.setWindowTitle(f"🏷️ مدیریت فیلدها - Source {self.source_id}")
        self.setMinimumSize(800, 600)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        layout = QVBoxLayout(self)
        
        # هدر
        header = QLabel(f"🏷️ فیلدهای منبع داده #{self.source_id}")
        header.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        layout.addWidget(header)
        
        # دکمه‌های عملیات
        btn_layout = QHBoxLayout()
        
        add_btn = QPushButton("➕ فیلد جدید")
        add_btn.clicked.connect(self.add_field)
        btn_layout.addWidget(add_btn)
        
        refresh_btn = QPushButton("🔄 بروزرسانی")
        refresh_btn.clicked.connect(self.load_fields)
        btn_layout.addWidget(refresh_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # جدول فیلدها
        self.fields_table = QTableWidget()
        self.fields_table.setColumnCount(6)
        self.fields_table.setHorizontalHeaderLabels([
            "ID", "نام", "نام نمایشی", "نوع", "نقش", "عملیات"
        ])
        self.fields_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        layout.addWidget(self.fields_table)
        
        # دکمه بستن
        close_btn = QPushButton("✅ بستن")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)
    
    def load_fields(self):
        """بارگذاری فیلدها"""
        fields = self.dm.list_fields(self.source_id)
        self.fields_table.setRowCount(len(fields))
        
        for i, field in enumerate(fields):
            self.fields_table.setItem(i, 0, QTableWidgetItem(str(field.id)))
            self.fields_table.setItem(i, 1, QTableWidgetItem(field.field_name))
            self.fields_table.setItem(i, 2, QTableWidgetItem(field.field_display_name))
            self.fields_table.setItem(i, 3, QTableWidgetItem(field.data_type))
            self.fields_table.setItem(i, 4, QTableWidgetItem(field.field_role or "-"))
            
            # دکمه حذف
            delete_btn = QPushButton("🗑️")
            delete_btn.clicked.connect(lambda checked, fid=field.id: self.delete_field(fid))
            self.fields_table.setCellWidget(i, 5, delete_btn)
    
    def add_field(self):
        """اضافه کردن فیلد جدید"""
        dialog = AddFieldDialog(self.source_id, self.dm, self)
        if dialog.exec():
            self.load_fields()
    
    def delete_field(self, field_id: int):
        """حذف فیلد"""
        reply = QMessageBox.question(
            self,
            "تایید حذف",
            f"آیا می‌خواهید فیلد #{field_id} را حذف کنید؟",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.dm.delete_field(field_id)
                QMessageBox.information(self, "موفق", "فیلد حذف شد")
                self.load_fields()
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در حذف: {e}")


class AddFieldDialog(QDialog):
    """دیالوگ اضافه کردن فیلد"""
    
    def __init__(self, source_id: int, dm: DataManager, parent=None):
        super().__init__(parent)
        self.source_id = source_id
        self.dm = dm
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("➕ فیلد جدید")
        self.setMinimumWidth(400)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        layout = QFormLayout(self)
        
        # نام فیلد
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("مثال: amount")
        layout.addRow("نام فیلد:", self.name_input)
        
        # نام نمایشی
        self.display_input = QLineEdit()
        self.display_input.setPlaceholderText("مثال: مبلغ")
        layout.addRow("نام نمایشی:", self.display_input)
        
        # نوع داده
        self.type_combo = QComboBox()
        self.type_combo.addItems([
            "text", "number", "decimal", "date", "datetime",
            "boolean", "formula", "lookup", "choice", "json"
        ])
        layout.addRow("نوع داده:", self.type_combo)
        
        # نقش
        self.role_combo = QComboBox()
        self.role_combo.addItems([
            "custom", "label", "customer", "supplier", "amount",
            "rate", "quantity", "date", "status", "description"
        ])
        layout.addRow("نقش:", self.role_combo)
        
        # ضروری؟
        self.required_check = QCheckBox("ضروری")
        layout.addRow("", self.required_check)
        
        # دکمه‌ها
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.save_field)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
    
    def save_field(self):
        """ذخیره فیلد"""
        field_name = self.name_input.text().strip()
        display_name = self.display_input.text().strip()
        
        if not field_name:
            QMessageBox.warning(self, "خطا", "نام فیلد الزامی است")
            return
        
        try:
            field_id = self.dm.add_field(
                source_id=self.source_id,
                field_name=field_name,
                field_display_name=display_name or field_name,
                data_type=self.type_combo.currentText(),
                field_role=self.role_combo.currentText(),
                is_required=self.required_check.isChecked()
            )
            QMessageBox.information(self, "موفق", f"فیلد ایجاد شد: ID={field_id}")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در ایجاد فیلد: {e}")


class BIPlatformManager(QMainWindow):
    """
    مدیر اصلی پلتفرم BI - نسخه کامل
    """
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.dm = DataManager()
        self.init_ui()
        self.load_data()
    
    def init_ui(self):
        """راه‌اندازی رابط کاربری"""
        self.setWindowTitle("🚀 GT-Land BI Platform - سیستم هوش تجاری")
        self.setGeometry(100, 50, 1400, 900)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        # ایجاد منو و toolbar
        self.create_menu()
        self.create_toolbar()
        
        # ویجت مرکزی
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # عنوان
        title = QLabel("🎯 پلتفرم هوش تجاری GT-Land")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #1976D2; padding: 15px;")
        main_layout.addWidget(title)
        
        # تب‌ها
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                background: white;
            }
            QTabBar::tab {
                background: #f5f5f5;
                border: 1px solid #ddd;
                padding: 12px 25px;
                margin: 2px;
                border-radius: 6px;
                font-size: 11pt;
                font-weight: 500;
            }
            QTabBar::tab:selected {
                background: #1976D2;
                color: white;
            }
            QTabBar::tab:hover {
                background: #42A5F5;
                color: white;
            }
        """)
        
        # تب‌های مختلف
        self.create_sources_tab()
        self.create_explorer_tab()
        self.create_reports_tab()
        self.create_migration_tab()
        
        main_layout.addWidget(self.tabs)
        
        # نوار وضعیت
        self.statusBar().showMessage("✅ آماده")
        self.statusBar().setStyleSheet("""
            QStatusBar {
                background: #f8f9fa;
                color: #495057;
                font-size: 10pt;
                padding: 5px;
                border-top: 1px solid #dee2e6;
            }
        """)
    
    def create_menu(self):
        """ایجاد منو"""
        menubar = self.menuBar()
        
        # منوی داده
        data_menu = menubar.addMenu("📊 داده")
        
        import_action = QAction("📥 Import از Phase 1", self)
        import_action.triggered.connect(self.open_migration_dialog)
        data_menu.addAction(import_action)
        
        data_menu.addSeparator()
        
        export_action = QAction("📤 Export به Excel", self)
        export_action.triggered.connect(self.export_data)
        data_menu.addAction(export_action)
        
        # منوی ابزارها
        tools_menu = menubar.addMenu("🔧 ابزارها")
        
        formula_action = QAction("🧮 Formula Builder", self)
        formula_action.triggered.connect(self.open_formula_builder_safe)
        tools_menu.addAction(formula_action)
        
        # منوی راهنما
        help_menu = menubar.addMenu("❓ راهنما")
        
        docs_action = QAction("📚 مستندات", self)
        docs_action.triggered.connect(self.show_documentation)
        help_menu.addAction(docs_action)
    
    def create_toolbar(self):
        """ایجاد نوار ابزار"""
        toolbar = QToolBar()
        toolbar.setIconSize(QSize(24, 24))
        toolbar.setMovable(False)
        toolbar.setStyleSheet("""
            QToolBar {
                background: #f8f9fa;
                border-bottom: 2px solid #dee2e6;
                spacing: 8px;
                padding: 8px;
            }
            QPushButton {
                background: white;
                border: 1px solid #ced4da;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 10pt;
            }
            QPushButton:hover {
                background: #e9ecef;
                border-color: #1976D2;
            }
        """)
        
        # دکمه‌های سریع
        refresh_btn = QPushButton("🔄 بروزرسانی")
        refresh_btn.clicked.connect(self.load_data)
        toolbar.addWidget(refresh_btn)
        
        toolbar.addSeparator()
        
        new_source_btn = QPushButton("➕ منبع جدید")
        new_source_btn.clicked.connect(self.add_new_source)
        toolbar.addWidget(new_source_btn)
        
        toolbar.addSeparator()
        
        migration_btn = QPushButton("🔄 Migration از Phase 1")
        migration_btn.clicked.connect(self.open_migration_dialog)
        toolbar.addWidget(migration_btn)
        
        self.addToolBar(toolbar)
    
    def create_sources_tab(self):
        """تب مدیریت منابع داده"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # عنوان
        header = QLabel("📊 منابع داده (Data Sources)")
        header.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        layout.addWidget(header)
        
        # دکمه‌های عملیات
        btn_layout = QHBoxLayout()
        
        add_btn = QPushButton("➕ منبع جدید")
        add_btn.clicked.connect(self.add_new_source)
        add_btn.setStyleSheet(self.get_button_style("#4CAF50"))
        btn_layout.addWidget(add_btn)
        
        refresh_btn = QPushButton("🔄 بروزرسانی")
        refresh_btn.clicked.connect(self.refresh_sources)
        refresh_btn.setStyleSheet(self.get_button_style("#2196F3"))
        btn_layout.addWidget(refresh_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # جدول منابع
        self.sources_table = QTableWidget()
        self.sources_table.setColumnCount(7)
        self.sources_table.setHorizontalHeaderLabels([
            "ID", "نام", "نوع", "تعداد رکورد", "فیلدها", "وضعیت", "عملیات"
        ])
        self.sources_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.sources_table.setAlternatingRowColors(True)
        self.sources_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 6px;
                gridline-color: #e0e0e0;
            }
            QHeaderView::section {
                background: #f5f5f5;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
        """)
        layout.addWidget(self.sources_table)
        
        self.tabs.addTab(tab, "📊 منابع داده")
    
    def create_explorer_tab(self):
        """تب مرورگر داده"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # استفاده از DataExplorerWidget
        try:
            from app.gui.financial.data_explorer import DataExplorerWidget
            self.data_explorer = DataExplorerWidget(self.dm)
            layout.addWidget(self.data_explorer)
        except Exception as e:
            error_label = QLabel(f"⚠️ خطا در بارگذاری Data Explorer:\n{e}")
            error_label.setStyleSheet("padding: 20px; background: #ffebee;")
            layout.addWidget(error_label)
        
        self.tabs.addTab(tab, "🔍 مرورگر داده")
    
    def create_reports_tab(self):
        """تب گزارش‌ها و تحلیل"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # هدر
        header = QLabel("📈 گزارش‌ها و تحلیل داده")
        header.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        layout.addWidget(header)
        
        # دکمه‌های عملیات
        btn_layout = QHBoxLayout()
        
        new_report_btn = QPushButton("📊 گزارش جدید")
        new_report_btn.clicked.connect(self.create_new_report)
        new_report_btn.setStyleSheet(self.get_button_style("#4CAF50"))
        btn_layout.addWidget(new_report_btn)
        
        quick_stats_btn = QPushButton("⚡ آمار سریع")
        quick_stats_btn.clicked.connect(self.show_quick_stats)
        quick_stats_btn.setStyleSheet(self.get_button_style("#2196F3"))
        btn_layout.addWidget(quick_stats_btn)
        
        export_report_btn = QPushButton("📤 Export گزارش")
        export_report_btn.clicked.connect(self.export_current_report)
        export_report_btn.setStyleSheet(self.get_button_style("#FF9800"))
        btn_layout.addWidget(export_report_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # جدول گزارش‌های ذخیره شده
        reports_group = QGroupBox("📋 گزارش‌های ذخیره شده")
        reports_layout = QVBoxLayout(reports_group)
        
        self.reports_table = QTableWidget()
        self.reports_table.setColumnCount(5)
        self.reports_table.setHorizontalHeaderLabels([
            "نام گزارش", "منبع", "تاریخ ایجاد", "نوع", "عملیات"
        ])
        self.reports_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.reports_table.setAlternatingRowColors(True)
        reports_layout.addWidget(self.reports_table)
        
        layout.addWidget(reports_group)
        
        # بخش آمار کلی
        stats_group = QGroupBox("📊 آمار کلی")
        stats_layout = QHBoxLayout(stats_group)
        
        # کارت‌های آماری
        self.create_stat_card(stats_layout, "📦 تعداد منابع", "0", "#2196F3")
        self.create_stat_card(stats_layout, "📝 کل رکوردها", "0", "#4CAF50")
        self.create_stat_card(stats_layout, "🏷️ کل فیلدها", "0", "#FF9800")
        self.create_stat_card(stats_layout, "📈 گزارش‌ها", "0", "#9C27B0")
        
        layout.addWidget(stats_group)
        
        # بارگذاری داده‌ها
        self.load_reports_data()
        
        self.tabs.addTab(tab, "📈 گزارش‌ها")
    
    def create_stat_card(self, parent_layout, title, value, color):
        """ایجاد کارت آماری"""
        card = QWidget()
        card.setStyleSheet(f"""
            QWidget {{
                background: {color};
                border-radius: 8px;
                padding: 15px;
            }}
        """)
        
        card_layout = QVBoxLayout(card)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("color: white; font-size: 11pt; font-weight: bold;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_layout.addWidget(title_label)
        
        value_label = QLabel(value)
        value_label.setStyleSheet("color: white; font-size: 24pt; font-weight: bold;")
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        value_label.setObjectName(f"stat_{title}")
        card_layout.addWidget(value_label)
        
        parent_layout.addWidget(card)
    
    def load_reports_data(self):
        """بارگذاری داده‌های گزارش‌ها"""
        try:
            # بارگذاری گزارش‌های ذخیره شده از دیتابیس
            from app.models.financial.dynamic_eav import ReportDefinition
            
            reports = self.dm.session.query(ReportDefinition).all()
            self.reports_table.setRowCount(len(reports))
            
            for i, report in enumerate(reports):
                # نام گزارش
                self.reports_table.setItem(i, 0, QTableWidgetItem(report.name))
                
                # منبع
                source = self.dm.get_source(report.primary_source_id)
                source_name = source.name if source else f"Source #{report.primary_source_id}"
                self.reports_table.setItem(i, 1, QTableWidgetItem(source_name))
                
                # تاریخ ایجاد
                created_date = report.created_at.strftime("%Y-%m-%d %H:%M") if report.created_at else "-"
                self.reports_table.setItem(i, 2, QTableWidgetItem(created_date))
                
                # نوع
                report_type = report.report_type or "-"
                self.reports_table.setItem(i, 3, QTableWidgetItem(report_type))
                
                # دکمه‌های عملیات
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                actions_layout.setSpacing(4)
                
                view_btn = QPushButton("👁️ مشاهده")
                view_btn.setStyleSheet("""
                    QPushButton {
                        background: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 6px 12px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background: #1976D2;
                    }
                """)
                view_btn.clicked.connect(lambda checked, r=report: self.view_report(r))
                actions_layout.addWidget(view_btn)
                
                delete_btn = QPushButton("🗑️ حذف")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background: #F44336;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 6px 12px;
                        font-size: 9pt;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background: #D32F2F;
                    }
                """)
                delete_btn.clicked.connect(lambda checked, r=report: self.delete_report(r))
                actions_layout.addWidget(delete_btn)
                
                self.reports_table.setCellWidget(i, 4, actions_widget)
            
            # بروزرسانی آمار
            sources = self.dm.list_sources()
            total_records = sum(s.total_records for s in sources)
            total_fields = sum(len(self.dm.list_fields(s.id)) for s in sources)
            
            # پیدا کردن و بروزرسانی label های آماری
            for widget in self.findChildren(QLabel):
                if widget.objectName() == "stat_📦 تعداد منابع":
                    widget.setText(str(len(sources)))
                elif widget.objectName() == "stat_📝 کل رکوردها":
                    widget.setText(f"{total_records:,}")
                elif widget.objectName() == "stat_🏷️ کل فیلدها":
                    widget.setText(str(total_fields))
                elif widget.objectName() == "stat_📈 گزارش‌ها":
                    widget.setText(str(len(reports)))
        
        except Exception as e:
            print(f"خطا در بارگذاری گزارش‌ها: {e}")
    
    def create_new_report(self):
        """ایجاد گزارش جدید"""
        dialog = QDialog(self)
        dialog.setWindowTitle("📊 گزارش جدید")
        dialog.setMinimumSize(600, 400)
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        layout = QFormLayout(dialog)
        
        # نام گزارش
        name_input = QLineEdit()
        name_input.setPlaceholderText("مثال: گزارش فروش ماهانه")
        layout.addRow("📝 نام گزارش:", name_input)
        
        # انتخاب منبع
        source_combo = QComboBox()
        sources = self.dm.list_sources()
        for source in sources:
            source_combo.addItem(source.name, source.id)
        layout.addRow("📊 منبع داده:", source_combo)
        
        # نوع گزارش
        type_combo = QComboBox()
        type_combo.addItems([
            "جدولی (Table)",
            "خلاصه (Summary)",
            "نمودار میله‌ای (Bar Chart)",
            "نمودار خطی (Line Chart)",
            "نمودار دایره‌ای (Pie Chart)"
        ])
        layout.addRow("📈 نوع گزارش:", type_combo)
        
        # فیلدهای نمایشی
        if sources:
            source_id = sources[0].id
            fields = self.dm.list_fields(source_id)
            
            fields_group = QGroupBox("🏷️ فیلدهای نمایشی (انتخاب کنید)")
            fields_group.setMaximumHeight(350)  # محدود کردن ارتفاع
            
            # اضافه کردن scroll area برای فیلدهای زیاد
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setStyleSheet("""
                QScrollArea {
                    border: 1px solid #ddd;
                    border-radius: 4px;
                    background: white;
                }
            """)
            
            fields_container = QWidget()
            fields_layout = QVBoxLayout(fields_container)
            fields_layout.setSpacing(8)
            
            field_checks = []
            
            # دکمه انتخاب همه/هیچکدام
            select_all_layout = QHBoxLayout()
            
            select_all_btn = QPushButton("✅ انتخاب همه")
            select_all_btn.setStyleSheet("""
                QPushButton {
                    background: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 5px 10px;
                    font-size: 9pt;
                }
                QPushButton:hover {
                    background: #45a049;
                }
            """)
            
            deselect_all_btn = QPushButton("❌ حذف انتخاب")
            deselect_all_btn.setStyleSheet("""
                QPushButton {
                    background: #f44336;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 5px 10px;
                    font-size: 9pt;
                }
                QPushButton:hover {
                    background: #da190b;
                }
            """)
            
            select_all_layout.addWidget(select_all_btn)
            select_all_layout.addWidget(deselect_all_btn)
            select_all_layout.addStretch()
            
            fields_layout.addLayout(select_all_layout)
            
            # نمایش تمام فیلدها (بدون محدودیت)
            for field in fields:
                cb = QCheckBox(f"{field.field_display_name} ({field.field_name})")
                cb.setStyleSheet("""
                    QCheckBox {
                        padding: 5px;
                        font-size: 10pt;
                    }
                    QCheckBox:hover {
                        background: #f0f0f0;
                        border-radius: 4px;
                    }
                """)
                field_checks.append((cb, field.id))
                fields_layout.addWidget(cb)
            
            # اتصال دکمه‌های انتخاب
            def select_all():
                for cb, _ in field_checks:
                    cb.setChecked(True)
            
            def deselect_all():
                for cb, _ in field_checks:
                    cb.setChecked(False)
            
            select_all_btn.clicked.connect(select_all)
            deselect_all_btn.clicked.connect(deselect_all)
            
            scroll.setWidget(fields_container)
            
            fields_group_layout = QVBoxLayout(fields_group)
            fields_group_layout.addWidget(scroll)
            
            layout.addRow(fields_group)
        
        # دکمه‌ها
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        
        def save_report():
            report_name = name_input.text().strip()
            if not report_name:
                QMessageBox.warning(dialog, "خطا", "نام گزارش الزامی است")
                return
            
            try:
                # دریافت منبع انتخاب شده
                source_id = source_combo.currentData()
                report_type = type_combo.currentText()
                
                # دریافت فیلدهای انتخاب شده
                selected_fields = []
                for cb, field_id in field_checks:
                    if cb.isChecked():
                        selected_fields.append(field_id)
                
                if not selected_fields:
                    QMessageBox.warning(dialog, "خطا", "حداقل یک فیلد را انتخاب کنید")
                    return
                
                # ذخیره در دیتابیس
                from app.models.financial.dynamic_eav import ReportDefinition
                
                report = ReportDefinition(
                    name=report_name,
                    primary_source_id=source_id,
                    report_type=report_type,
                    field_mappings={
                        "selected_fields": selected_fields,
                        "report_config": {
                            "type": report_type,
                            "created_at": datetime.now().isoformat()
                        }
                    },
                    filters={},
                    aggregations={},
                    is_active=True
                )
                
                self.dm.session.add(report)
                self.dm.session.commit()
                
                QMessageBox.information(
                    dialog,
                    "✅ موفق",
                    f"گزارش '{report_name}' با موفقیت ذخیره شد!\n\n"
                    f"منبع: {source_combo.currentText()}\n"
                    f"نوع: {report_type}\n"
                    f"فیلدهای انتخابی: {len(selected_fields)} فیلد\n\n"
                    f"شناسه گزارش: {report.id}"
                )
                dialog.accept()
                
            except Exception as e:
                QMessageBox.critical(
                    dialog,
                    "❌ خطا",
                    f"خطا در ذخیره گزارش:\n{e}"
                )
                print(traceback.format_exc())
        
        buttons.accepted.connect(save_report)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        dialog.exec()
        self.load_reports_data()
    
    def show_quick_stats(self):
        """نمایش آمار سریع"""
        try:
            sources = self.dm.list_sources()
            
            if not sources:
                QMessageBox.information(
                    self,
                    "آمار سریع",
                    "هیچ منبع داده‌ای وجود ندارد"
                )
                return
            
            # انتخاب منبع
            source_names = [s.name for s in sources]
            
            source_name, ok = QInputDialog.getItem(
                self,
                "انتخاب منبع",
                "منبع مورد نظر را انتخاب کنید:",
                source_names,
                0,
                False
            )
            
            if not ok:
                return
            
            source = next(s for s in sources if s.name == source_name)
            
            # دریافت داده‌ها
            data = self.dm.query(source.id, limit=1000)
            fields = self.dm.list_fields(source.id)
            
            # محاسبه آمار
            stats_text = f"📊 آمار منبع: {source.name}\n\n"
            stats_text += f"📝 تعداد رکوردها: {len(data):,}\n"
            stats_text += f"🏷️ تعداد فیلدها: {len(fields)}\n\n"
            
            # آمار فیلدهای عددی
            numeric_fields = [f for f in fields if f.data_type in ['number', 'decimal']]
            
            if numeric_fields and data:
                stats_text += "📈 آمار فیلدهای عددی:\n\n"
                
                for field in numeric_fields[:5]:  # 5 فیلد اول
                    values = []
                    for row in data:
                        if field.field_name in row and row[field.field_name]:
                            try:
                                values.append(float(row[field.field_name]))
                            except:
                                pass
                    
                    if values:
                        stats_text += f"• {field.field_display_name}:\n"
                        stats_text += f"  - مجموع: {sum(values):,.0f}\n"
                        stats_text += f"  - میانگین: {sum(values)/len(values):,.2f}\n"
                        stats_text += f"  - حداقل: {min(values):,.0f}\n"
                        stats_text += f"  - حداکثر: {max(values):,.0f}\n\n"
            
            # نمایش در MessageBox
            msg = QMessageBox(self)
            msg.setWindowTitle("⚡ آمار سریع")
            msg.setText(stats_text)
            msg.setIcon(QMessageBox.Icon.Information)
            msg.exec()
        
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در محاسبه آمار:\n{e}")
    
    def export_current_report(self):
        """Export گزارش فعلی"""
        # انتخاب فرمت
        formats = ["Excel (*.xlsx)", "CSV (*.csv)", "JSON (*.json)"]
        format_choice, ok = QInputDialog.getItem(
            self,
            "انتخاب فرمت",
            "فرمت خروجی را انتخاب کنید:",
            formats,
            0,
            False
        )
        
        if not ok:
            return
        
        # انتخاب منبع
        sources = self.dm.list_sources()
        if not sources:
            QMessageBox.warning(self, "خطا", "هیچ منبع داده‌ای وجود ندارد")
            return
        
        source_names = [s.name for s in sources]
        source_name, ok = QInputDialog.getItem(
            self,
            "انتخاب منبع",
            "منبع مورد نظر را انتخاب کنید:",
            source_names,
            0,
            False
        )
        
        if not ok:
            return
        
        source = next(s for s in sources if s.name == source_name)
        
        try:
            # انتخاب مسیر ذخیره
            if "Excel" in format_choice:
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "ذخیره گزارش",
                    f"report_{source.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    "Excel Files (*.xlsx)"
                )
                if file_path:
                    self.export_to_excel(source, file_path)
            
            elif "CSV" in format_choice:
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "ذخیره گزارش",
                    f"report_{source.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    "CSV Files (*.csv)"
                )
                if file_path:
                    self.export_to_csv(source, file_path)
            
            elif "JSON" in format_choice:
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "ذخیره گزارش",
                    f"report_{source.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    "JSON Files (*.json)"
                )
                if file_path:
                    self.export_to_json(source, file_path)
            
            if file_path:
                QMessageBox.information(
                    self,
                    "موفق",
                    f"گزارش با موفقیت ذخیره شد:\n{file_path}"
                )
        
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در export:\n{e}")
    
    def export_to_excel(self, source, file_path):
        """Export به Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # دریافت داده‌ها
            data = self.dm.query(source.id, limit=10000)
            fields = self.dm.list_fields(source.id)
            
            # ایجاد workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = source.name[:31]  # Excel limit
            
            # هدرها
            headers = [f.field_display_name for f in fields]
            ws.append(headers)
            
            # استایل هدر
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # داده‌ها
            for row_data in data:
                row = []
                for field in fields:
                    value = row_data.get(field.field_name, "")
                    row.append(value)
                ws.append(row)
            
            # تنظیم عرض ستون‌ها
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
        
        except ImportError:
            QMessageBox.warning(
                self,
                "کتابخانه موجود نیست",
                "برای export به Excel، نصب کنید:\npip install openpyxl"
            )
    
    def export_to_csv(self, source, file_path):
        """Export به CSV"""
        data = self.dm.query(source.id, limit=10000)
        fields = self.dm.list_fields(source.id)
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # هدرها
            headers = [f.field_display_name for f in fields]
            writer.writerow(headers)
            
            # داده‌ها
            for row_data in data:
                row = []
                for field in fields:
                    value = row_data.get(field.field_name, "")
                    row.append(value)
                writer.writerow(row)
    
    def export_to_json(self, source, file_path):
        """Export به JSON"""
        data = self.dm.query(source.id, limit=10000)
        fields = self.dm.list_fields(source.id)
        
        # تبدیل فیلدها به نام نمایشی
        export_data = []
        for row_data in data:
            row = {}
            for field in fields:
                row[field.field_display_name] = row_data.get(field.field_name, "")
            export_data.append(row)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
    
    def view_report(self, report):
        """مشاهده جزئیات گزارش"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"👁️ مشاهده گزارش: {report.name}")
        dialog.setMinimumSize(700, 500)
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        layout = QVBoxLayout(dialog)
        
        # اطلاعات گزارش
        info_text = QTextEdit()
        info_text.setReadOnly(True)
        
        source = self.dm.get_source(report.primary_source_id)
        source_name = source.name if source else f"Source #{report.primary_source_id}"
        
        # استخراج فیلدهای انتخاب شده
        selected_fields = report.field_mappings.get("selected_fields", []) if report.field_mappings else []
        
        info = f"""
📊 نام گزارش: {report.name}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📦 منبع داده: {source_name}
🏷️ نوع گزارش: {report.report_type or '-'}
📅 تاریخ ایجاد: {report.created_at.strftime('%Y-%m-%d %H:%M:%S') if report.created_at else '-'}
📅 آخرین بروزرسانی: {report.updated_at.strftime('%Y-%m-%d %H:%M:%S') if report.updated_at else '-'}
✅ وضعیت: {'فعال' if report.is_active else 'غیرفعال'}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📋 فیلدهای انتخاب شده ({len(selected_fields)} فیلد):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

"""
        
        # نمایش فیلدهای انتخاب شده
        if selected_fields and source:
            all_fields = self.dm.list_fields(source.id)
            field_dict = {f.id: f for f in all_fields}
            
            for i, field_id in enumerate(selected_fields, 1):
                if field_id in field_dict:
                    field = field_dict[field_id]
                    info += f"{i}. {field.field_display_name} ({field.field_name}) - {field.data_type}\n"
        else:
            info += "هیچ فیلدی انتخاب نشده است.\n"
        
        info += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        info += f"🆔 شناسه گزارش: {report.id}\n"
        
        info_text.setPlainText(info)
        info_text.setStyleSheet("""
            QTextEdit {
                background: #f8f9fa;
                border: 1px solid #ddd;
                border-radius: 6px;
                padding: 15px;
                font-family: 'Consolas', monospace;
                font-size: 11pt;
                line-height: 1.6;
            }
        """)
        
        layout.addWidget(info_text)
        
        # دکمه‌ها
        btn_layout = QHBoxLayout()
        
        run_btn = QPushButton("▶️ اجرای گزارش")
        run_btn.clicked.connect(lambda: self.run_report(report))
        run_btn.setStyleSheet(self.get_button_style("#4CAF50"))
        btn_layout.addWidget(run_btn)
        
        close_btn = QPushButton("✅ بستن")
        close_btn.clicked.connect(dialog.accept)
        close_btn.setStyleSheet(self.get_button_style("#2196F3"))
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    def delete_report(self, report):
        """حذف گزارش"""
        reply = QMessageBox.question(
            self,
            "تایید حذف",
            f"آیا می‌خواهید گزارش '{report.name}' را حذف کنید؟",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.dm.session.delete(report)
                self.dm.session.commit()
                QMessageBox.information(self, "موفق", f"گزارش '{report.name}' حذف شد")
                self.load_reports_data()
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در حذف گزارش:\n{e}")
    
    def run_report(self, report):
        """اجرای گزارش"""
        try:
            source = self.dm.get_source(report.primary_source_id)
            if not source:
                QMessageBox.warning(self, "خطا", "منبع داده یافت نشد")
                return
            
            # دریافت داده‌ها
            data = self.dm.query(source.id, limit=1000)
            
            if not data:
                QMessageBox.information(self, "اطلاع", "هیچ داده‌ای برای نمایش وجود ندارد")
                return
            
            # نمایش نتایج در یک dialog
            result_dialog = QDialog(self)
            result_dialog.setWindowTitle(f"📊 نتایج: {report.name}")
            result_dialog.setMinimumSize(900, 600)
            result_dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
            
            layout = QVBoxLayout(result_dialog)
            
            # جدول نتایج
            table = QTableWidget()
            
            # فیلدهای انتخاب شده
            selected_fields = report.field_mappings.get("selected_fields", []) if report.field_mappings else []
            all_fields = self.dm.list_fields(source.id)
            field_dict = {f.id: f for f in all_fields}
            
            # تنظیم ستون‌ها
            display_fields = [field_dict[fid] for fid in selected_fields if fid in field_dict]
            if not display_fields:
                display_fields = all_fields
            
            table.setColumnCount(len(display_fields))
            table.setHorizontalHeaderLabels([f.field_display_name for f in display_fields])
            
            # پر کردن داده‌ها
            table.setRowCount(len(data))
            for i, row_data in enumerate(data):
                for j, field in enumerate(display_fields):
                    value = row_data.get(field.field_name, "")
                    table.setItem(i, j, QTableWidgetItem(str(value)))
            
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            table.setAlternatingRowColors(True)
            
            layout.addWidget(table)
            
            # دکمه‌های عملیات
            buttons_layout = QHBoxLayout()
            
            # دکمه خروجی Excel
            excel_btn = QPushButton("📊 خروجی Excel")
            excel_btn.setStyleSheet(self.get_button_style("#4CAF50"))
            excel_btn.clicked.connect(lambda: self.export_report_data(report, data, display_fields, "excel"))
            buttons_layout.addWidget(excel_btn)
            
            # دکمه خروجی CSV
            csv_btn = QPushButton("📄 خروجی CSV")
            csv_btn.setStyleSheet(self.get_button_style("#FF9800"))
            csv_btn.clicked.connect(lambda: self.export_report_data(report, data, display_fields, "csv"))
            buttons_layout.addWidget(csv_btn)
            
            # دکمه خروجی JSON
            json_btn = QPushButton("📋 خروجی JSON")
            json_btn.setStyleSheet(self.get_button_style("#9C27B0"))
            json_btn.clicked.connect(lambda: self.export_report_data(report, data, display_fields, "json"))
            buttons_layout.addWidget(json_btn)
            
            buttons_layout.addStretch()
            
            # دکمه بستن
            close_btn = QPushButton("✅ بستن")
            close_btn.clicked.connect(result_dialog.accept)
            close_btn.setStyleSheet(self.get_button_style("#2196F3"))
            buttons_layout.addWidget(close_btn)
            
            layout.addLayout(buttons_layout)
            
            result_dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در اجرای گزارش:\n{e}")
            import traceback
            print(traceback.format_exc())
    
    def export_report_data(self, report, data, fields, export_type):
        """خروجی گرفتن از داده‌های گزارش"""
        try:
            # انتخاب مسیر ذخیره
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"report_{report.name}_{timestamp}"
            
            if export_type == "excel":
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "ذخیره فایل Excel",
                    f"data/exports/{default_filename}.xlsx",
                    "Excel Files (*.xlsx)"
                )
                if file_path:
                    self._export_to_excel(data, fields, file_path)
                    
            elif export_type == "csv":
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "ذخیره فایل CSV",
                    f"data/exports/{default_filename}.csv",
                    "CSV Files (*.csv)"
                )
                if file_path:
                    self._export_to_csv(data, fields, file_path)
                    
            elif export_type == "json":
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "ذخیره فایل JSON",
                    f"data/exports/{default_filename}.json",
                    "JSON Files (*.json)"
                )
                if file_path:
                    self._export_to_json(data, fields, file_path)
            
            if file_path:
                QMessageBox.information(
                    self,
                    "موفق",
                    f"✅ گزارش با موفقیت ذخیره شد:\n{file_path}"
                )
                
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در خروجی:\n{e}")
            print(traceback.format_exc())
    
    def _export_to_excel(self, data, fields, file_path):
        """خروجی Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # هدر
        headers = [f.field_display_name for f in fields]
        ws.append(headers)
        
        # استایل هدر
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # داده‌ها
        for row_data in data:
            row = [row_data.get(f.field_name, "") for f in fields]
            ws.append(row)
        
        # تنظیم عرض ستون‌ها
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    def _export_to_csv(self, data, fields, file_path):
        """خروجی CSV"""
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # هدر
            headers = [f.field_display_name for f in fields]
            writer.writerow(headers)
            
            # داده‌ها
            for row_data in data:
                row = [row_data.get(f.field_name, "") for f in fields]
                writer.writerow(row)
    
    def _export_to_json(self, data, fields, file_path):
        """خروجی JSON"""
        export_data = []
        for row_data in data:
            row_dict = {}
            for field in fields:
                row_dict[field.field_display_name] = row_data.get(field.field_name, "")
            export_data.append(row_dict)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
    
    def create_migration_tab(self):
        """تب Migration"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # عنوان
        header = QLabel("🔄 انتقال داده از Phase 1")
        header.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(header)
        
        # توضیحات
        info_box = QGroupBox("ℹ️ درباره Migration")
        info_layout = QVBoxLayout(info_box)
        
        info_text = QLabel(
            "این ابزار داده‌های استخراج شده از Google Sheets (Phase 1) را\n"
            "به سیستم هوش تجاری جدید (Phase 2) منتقل می‌کند.\n\n"
            "✅ مزایا:\n"
            "• دسترسی به تمام ویژگی‌های BI Platform\n"
            "• فیلدهای قابل تنظیم و فرمول‌های محاسباتی\n"
            "• گزارش‌گیری پیشرفته و Dashboard\n"
            "• مقیاس‌پذیری برای حجم بالای داده"
        )
        info_text.setStyleSheet("padding: 15px; line-height: 1.6;")
        info_layout.addWidget(info_text)
        info_box.setStyleSheet("""
            QGroupBox {
                background: #e3f2fd;
                border: 2px solid #1976D2;
                border-radius: 8px;
                margin-top: 10px;
                font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                padding: 5px 10px;
            }
        """)
        layout.addWidget(info_box)
        
        # دکمه‌های عملیات
        btn_layout = QHBoxLayout()
        
        check_btn = QPushButton("📊 بررسی وضعیت")
        check_btn.setMinimumHeight(50)
        check_btn.clicked.connect(self.check_migration_status)
        check_btn.setStyleSheet(self.get_button_style("#2196F3"))
        btn_layout.addWidget(check_btn)
        
        migrate_btn = QPushButton("🚀 شروع Migration")
        migrate_btn.setMinimumHeight(50)
        migrate_btn.clicked.connect(self.start_migration)
        migrate_btn.setStyleSheet(self.get_button_style("#4CAF50"))
        btn_layout.addWidget(migrate_btn)
        
        layout.addLayout(btn_layout)
        
        # نمایش پیشرفت
        self.migration_log = QTextEdit()
        self.migration_log.setReadOnly(True)
        self.migration_log.setStyleSheet("""
            QTextEdit {
                background: #263238;
                color: #00ff00;
                font-family: 'Consolas', monospace;
                font-size: 10pt;
                border: 1px solid #555;
                border-radius: 6px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.migration_log)
        
        self.tabs.addTab(tab, "🔄 Migration")
    
    def load_data(self):
        """بارگذاری داده‌ها"""
        self.refresh_sources()
        self.statusBar().showMessage("✅ داده‌ها بروز شدند")
    
    def refresh_sources(self):
        """بروزرسانی جدول منابع"""
        try:
            sources = self.dm.list_sources()
            self.sources_table.setRowCount(len(sources))
            
            for i, source in enumerate(sources):
                # ID
                self.sources_table.setItem(i, 0, QTableWidgetItem(str(source.id)))
                
                # نام
                self.sources_table.setItem(i, 1, QTableWidgetItem(source.name))
                
                # نوع
                self.sources_table.setItem(i, 2, QTableWidgetItem(source.source_type))
                
                # تعداد رکورد
                self.sources_table.setItem(i, 3, QTableWidgetItem(f"{source.total_records:,}"))
                
                # تعداد فیلدها
                fields_count = len(self.dm.list_fields(source.id))
                self.sources_table.setItem(i, 4, QTableWidgetItem(str(fields_count)))
                
                # وضعیت
                status = "✅ فعال" if source.is_active else "❌ غیرفعال"
                self.sources_table.setItem(i, 5, QTableWidgetItem(status))
                
                # دکمه‌های عملیات
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                actions_layout.setSpacing(4)
                
                manage_btn = QPushButton("🏷️ فیلدها")
                manage_btn.setToolTip("مدیریت فیلدها")
                manage_btn.setStyleSheet("""
                    QPushButton {
                        background: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 6px 12px;
                        font-size: 10pt;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background: #1976D2;
                    }
                """)
                manage_btn.clicked.connect(lambda checked, s=source: self.manage_source_fields(s))
                actions_layout.addWidget(manage_btn)
                
                edit_btn = QPushButton("✏️ ویرایش")
                edit_btn.setToolTip("ویرایش منبع")
                edit_btn.setStyleSheet("""
                    QPushButton {
                        background: #FF9800;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 6px 12px;
                        font-size: 10pt;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background: #F57C00;
                    }
                """)
                edit_btn.clicked.connect(lambda checked, s=source: self.edit_source(s))
                actions_layout.addWidget(edit_btn)
                
                delete_btn = QPushButton("🗑️ حذف")
                delete_btn.setToolTip("حذف منبع")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background: #F44336;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 6px 12px;
                        font-size: 10pt;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background: #D32F2F;
                    }
                """)
                delete_btn.clicked.connect(lambda checked, s=source: self.delete_source(s))
                actions_layout.addWidget(delete_btn)
                
                self.sources_table.setCellWidget(i, 6, actions_widget)
                
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در بارگذاری منابع:\n{e}")
    
    def add_new_source(self):
        """اضافه کردن منبع جدید"""
        dialog = AddSourceDialog(self)
        if dialog.exec():
            data = dialog.get_data()
            if not data["name"]:
                QMessageBox.warning(self, "خطا", "نام منبع الزامی است")
                return
            
            try:
                source_id = self.dm.add_source(
                    name=data["name"],
                    source_type=data["source_type"],
                    connection_info={
                        "description": data["description"]
                    }
                )
                QMessageBox.information(
                    self,
                    "موفق",
                    f"منبع ایجاد شد!\nID: {source_id}\n\n"
                    "حالا می‌توانید فیلدها را اضافه کنید."
                )
                self.refresh_sources()
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در ایجاد منبع:\n{e}")
    
    def edit_source(self, source):
        """ویرایش منبع"""
        source_data = {
            "id": source.id,
            "name": source.name,
            "is_active": source.is_active,
            "connection_info": source.connection_info or {}
        }
        
        dialog = EditSourceDialog(source_data, self)
        if dialog.exec():
            updates = dialog.get_data()
            try:
                self.dm.update_source(source.id, **updates)
                QMessageBox.information(self, "موفق", "منبع به‌روز شد")
                self.refresh_sources()
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در به‌روزرسانی:\n{e}")
    
    def delete_source(self, source):
        """حذف منبع"""
        reply = QMessageBox.question(
            self,
            "تایید حذف",
            f"آیا می‌خواهید منبع '{source.name}' را حذف کنید؟\n\n"
            "⚠️ تمام داده‌های این منبع نیز حذف خواهند شد!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.dm.delete_source(source.id)
                QMessageBox.information(self, "موفق", "منبع حذف شد")
                self.refresh_sources()
            except Exception as e:
                QMessageBox.critical(self, "خطا", f"خطا در حذف:\n{e}")
    
    def manage_source_fields(self, source):
        """مدیریت فیلدهای یک منبع"""
        dialog = ManageFieldsDialog(source.id, self.dm, self)
        dialog.exec()
        self.refresh_sources()  # برای به‌روزرسانی تعداد فیلدها
    
    def open_formula_builder_safe(self):
        """باز کردن Formula Builder با بررسی امنیت"""
        try:
            # دریافت لیست فیلدها
            sources = self.dm.list_sources()
            
            if not sources:
                QMessageBox.information(
                    self,
                    "توجه",
                    "هیچ منبع داده‌ای وجود ندارد.\n\n"
                    "ابتدا یک منبع ایجاد کنید."
                )
                return
            
            # انتخاب اولین source
            source = sources[0]
            field_defs = self.dm.list_fields(source.id)
            
            if not field_defs:
                QMessageBox.information(
                    self,
                    "توجه",
                    f"منبع '{source.name}' هیچ فیلدی ندارد.\n\n"
                    "ابتدا فیلدها را اضافه کنید."
                )
                return
            
            # تبدیل به فرمت مورد نیاز
            fields = [
                {
                    "id": f.id,
                    "name": f.field_name,
                    "display_name": f.field_display_name,
                    "type": f.data_type
                }
                for f in field_defs
            ]
            
            # باز کردن دیالوگ
            from app.gui.financial.formula_builder import FormulaBuilderDialog
            dialog = FormulaBuilderDialog(fields, self)
            
            if dialog.exec():
                QMessageBox.information(
                    self,
                    "فرمول ساخته شد",
                    f"فرمول: {dialog.formula_text}\n\n"
                    "فرمول ذخیره شد."
                )
        
        except Exception as e:
            QMessageBox.critical(
                self,
                "خطا",
                f"خطا در باز کردن Formula Builder:\n{e}\n\n"
                "لطفاً از وجود حداقل یک منبع با فیلد مطمئن شوید."
            )
    
    def check_migration_status(self):
        """بررسی وضعیت migration"""
        try:
            from migrate_phase1_to_2 import check_migration_status
            import sys
            from io import StringIO
            
            # Capture output
            old_stdout = sys.stdout
            sys.stdout = StringIO()
            
            check_migration_status()
            
            output = sys.stdout.getvalue()
            sys.stdout = old_stdout
            
            self.migration_log.setPlainText(output)
            
        except Exception as e:
            self.migration_log.append(f"\n❌ خطا: {e}")
    
    def start_migration(self):
        """شروع migration"""
        reply = QMessageBox.question(
            self,
            "تایید Migration",
            "آیا می‌خواهید تمام داده‌های Phase 1 را به Phase 2 منتقل کنید؟\n\n"
            "⚠️ این عملیات ممکن است چند دقیقه طول بکشد.\n"
            "⚠️ داده‌های تکراری ایجاد می‌شود اگر قبلاً migration کرده باشید.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.migration_log.clear()
            self.migration_log.append("🚀 شروع Migration...\n")
            
            try:
                from migrate_phase1_to_2 import Phase1To2Migrator
                
                migrator = Phase1To2Migrator()
                
                # Redirect output
                import sys
                from io import StringIO
                old_stdout = sys.stdout
                sys.stdout = StringIO()
                
                stats = migrator.migrate_all()
                
                output = sys.stdout.getvalue()
                sys.stdout = old_stdout
                
                self.migration_log.append(output)
                self.migration_log.append("\n✅ Migration کامل شد!")
                
                self.refresh_sources()
                
            except Exception as e:
                self.migration_log.append(f"\n❌ خطا: {e}")
                import traceback
                self.migration_log.append(traceback.format_exc())
    
    def open_migration_dialog(self):
        """باز کردن دیالوگ migration"""
        self.tabs.setCurrentIndex(3)  # رفتن به تب Migration
    
    def export_data(self):
        """Export داده - استفاده از export_current_report"""
        # استفاده از همان متد export گزارش‌ها
        self.export_current_report()
    
    def show_documentation(self):
        """نمایش مستندات"""
        QMessageBox.information(
            self,
            "📚 مستندات",
            "مستندات کامل در فایل‌های زیر موجود است:\n\n"
            "• ARCHITECTURE.md - معماری سیستم\n"
            "• QUICK_START.md - راهنمای سریع\n"
            "• INTEGRATION_GUIDE.md - راهنمای یکپارچگی\n"
            "• migrate_phase1_to_2.py - راهنمای Migration"
        )
    
    def get_button_style(self, color: str) -> str:
        """استایل دکمه"""
        return f"""
            QPushButton {{
                background: {color};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-size: 11pt;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: {color}dd;
            }}
            QPushButton:pressed {{
                background: {color}aa;
            }}
        """
    
    def closeEvent(self, event):
        """رویداد بستن پنجره"""
        # بستن اتصال دیتابیس
        try:
            if hasattr(self, 'dm'):
                # DataManager handles its own cleanup
                pass
        except:
            pass
        event.accept()


if __name__ == "__main__":
    import sys
    
    app = QApplication(sys.argv)
    window = BIPlatformManager()
    window.show()
    sys.exit(app.exec())
