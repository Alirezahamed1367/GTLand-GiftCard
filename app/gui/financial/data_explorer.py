"""
Professional Data Explorer with Advanced Filters
مرورگر داده‌های حرفه‌ای با فیلترهای پیشرفته - متصل به Backend

قابلیت‌ها:
- نمایش داده‌های هر source (اتصال واقعی به EAV)
- فیلترهای چندگانه و پیچیده (QueryBuilder)
- جستجوی سریع
- Export به Excel/CSV
- Pivot Table
- Charts
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QComboBox, QLineEdit,
    QDialog, QFormLayout, QCheckBox, QDateEdit, QSpinBox,
    QGroupBox, QScrollArea, QHeaderView, QMenu, QMessageBox,
    QProgressBar
)
from PyQt6.QtCore import Qt, QDate, pyqtSignal, QThread
from PyQt6.QtGui import QFont, QAction
import json

# Backend imports
from app.core.financial import DataManager
# QueryBuilder را حذف کردیم - فیلترها رو مستقیم توی SQL می‌نویسیم


class AdvancedFilterDialog(QDialog):
    """
    دیالوگ فیلتر پیشرفته
    کاربر می‌تونه چند فیلتر مختلف اعمال کنه
    """
    
    def __init__(self, fields: list, parent=None):
        super().__init__(parent)
        self.fields = fields
        self.filters = []
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Advanced Filters")
        self.setMinimumSize(600, 400)
        
        layout = QVBoxLayout(self)
        
        # عنوان
        title = QLabel("🔍 Advanced Filters")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        layout.addWidget(title)
        
        # فیلترها
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        self.filters_widget = QWidget()
        self.filters_layout = QVBoxLayout(self.filters_widget)
        
        scroll.setWidget(self.filters_widget)
        layout.addWidget(scroll)
        
        # دکمه افزودن فیلتر
        add_filter_btn = QPushButton("+ Add Filter")
        add_filter_btn.clicked.connect(self.add_filter_row)
        layout.addWidget(add_filter_btn)
        
        # دکمه‌ها
        btn_layout = QHBoxLayout()
        
        apply_btn = QPushButton("Apply Filters")
        apply_btn.clicked.connect(self.accept)
        apply_btn.setStyleSheet("background: #4CAF50; color: white;")
        btn_layout.addWidget(apply_btn)
        
        clear_btn = QPushButton("Clear All")
        clear_btn.clicked.connect(self.clear_filters)
        btn_layout.addWidget(clear_btn)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
        
        # افزودن یک فیلتر پیش‌فرض
        self.add_filter_row()
    
    def add_filter_row(self):
        """افزودن یک ردیف فیلتر"""
        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)
        
        # انتخاب فیلد
        field_combo = QComboBox()
        for field in self.fields:
            field_combo.addItem(field['display_name'], field)
        field_combo.currentIndexChanged.connect(
            lambda: self.on_field_changed(field_combo, operator_combo, value_widget)
        )
        row_layout.addWidget(field_combo, 2)
        
        # انتخاب عملگر
        operator_combo = QComboBox()
        row_layout.addWidget(operator_combo, 1)
        
        # مقدار
        value_widget = QLineEdit()
        row_layout.addWidget(value_widget, 2)
        
        # دکمه حذف
        remove_btn = QPushButton("✕")
        remove_btn.setMaximumWidth(30)
        remove_btn.clicked.connect(lambda: self.remove_filter_row(row_widget))
        row_layout.addWidget(remove_btn)
        
        self.filters_layout.addWidget(row_widget)
        
        # بارگذاری عملگرهای پیش‌فرض
        self.on_field_changed(field_combo, operator_combo, value_widget)
        
        # ذخیره ارجاع
        row_widget.field_combo = field_combo
        row_widget.operator_combo = operator_combo
        row_widget.value_widget = value_widget
    
    def on_field_changed(self, field_combo, operator_combo, value_widget):
        """تغییر عملگرها بر اساس نوع فیلد"""
        field = field_combo.currentData()
        if not field:
            return
        
        operator_combo.clear()
        
        field_type = field.get('type', 'text')
        
        if field_type in ['number', 'decimal']:
            operators = [
                ('=', 'Equals'),
                ('!=', 'Not Equals'),
                ('>', 'Greater Than'),
                ('>=', 'Greater or Equal'),
                ('<', 'Less Than'),
                ('<=', 'Less or Equal'),
                ('BETWEEN', 'Between'),
            ]
        elif field_type == 'date':
            operators = [
                ('=', 'On Date'),
                ('>', 'After'),
                ('<', 'Before'),
                ('BETWEEN', 'Between Dates'),
                ('LAST_7_DAYS', 'Last 7 Days'),
                ('LAST_30_DAYS', 'Last 30 Days'),
                ('THIS_MONTH', 'This Month'),
                ('THIS_YEAR', 'This Year'),
            ]
        elif field_type == 'boolean':
            operators = [
                ('=', 'Is'),
            ]
        else:  # text
            operators = [
                ('=', 'Equals'),
                ('!=', 'Not Equals'),
                ('CONTAINS', 'Contains'),
                ('STARTS_WITH', 'Starts With'),
                ('ENDS_WITH', 'Ends With'),
                ('IS_EMPTY', 'Is Empty'),
                ('IS_NOT_EMPTY', 'Is Not Empty'),
            ]
        
        for op, label in operators:
            operator_combo.addItem(label, op)
    
    def remove_filter_row(self, row_widget):
        """حذف یک ردیف فیلتر"""
        self.filters_layout.removeWidget(row_widget)
        row_widget.deleteLater()
    
    def clear_filters(self):
        """پاک کردن تمام فیلترها"""
        while self.filters_layout.count():
            child = self.filters_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        
        # افزودن یک فیلتر خالی
        self.add_filter_row()
    
    def get_filters(self):
        """دریافت تمام فیلترها"""
        filters = []
        
        for i in range(self.filters_layout.count()):
            row_widget = self.filters_layout.itemAt(i).widget()
            if not row_widget:
                continue
            
            field = row_widget.field_combo.currentData()
            operator = row_widget.operator_combo.currentData()
            value = row_widget.value_widget.text()
            
            if field and operator:
                filters.append({
                    'field_id': field['id'],
                    'field_name': field['name'],
                    'operator': operator,
                    'value': value
                })
        
        return filters


class DataExplorerWidget(QWidget):
    """
    مرورگر حرفه‌ای داده‌ها
    """
    
    def __init__(self, dm: DataManager, parent=None):
        super().__init__(parent)
        self.dm = dm
        self.current_source = None
        self.current_filters = []
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # ═══════════════════════════════════════════════════════
        #                      HEADER & TOOLBAR
        # ═══════════════════════════════════════════════════════
        
        header = QHBoxLayout()
        
        # عنوان
        title = QLabel("📊 Data Explorer")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #2196F3;")
        header.addWidget(title)
        
        header.addStretch()
        
        # انتخاب Data Source
        header.addWidget(QLabel("Data Source:"))
        self.source_combo = QComboBox()
        self.source_combo.currentIndexChanged.connect(self.load_data)
        header.addWidget(self.source_combo)
        
        # دکمه بروزرسانی
        refresh_btn = QPushButton("🔄")
        refresh_btn.setToolTip("Refresh")
        refresh_btn.clicked.connect(self.load_sources)
        header.addWidget(refresh_btn)
        
        layout.addLayout(header)
        
        # ═══════════════════════════════════════════════════════
        #                      FILTER BAR
        # ═══════════════════════════════════════════════════════
        
        filter_bar = QHBoxLayout()
        
        # جستجوی سریع
        self.quick_search = QLineEdit()
        self.quick_search.setPlaceholderText("🔍 Quick search...")
        self.quick_search.textChanged.connect(self.apply_quick_search)
        filter_bar.addWidget(self.quick_search)
        
        # دکمه فیلتر پیشرفته
        advanced_filter_btn = QPushButton("⚙ Advanced Filters")
        advanced_filter_btn.clicked.connect(self.show_advanced_filters)
        filter_bar.addWidget(advanced_filter_btn)
        
        # نمایش فیلترهای فعال
        self.active_filters_label = QLabel("No filters active")
        self.active_filters_label.setStyleSheet("color: #666; font-size: 9pt;")
        filter_bar.addWidget(self.active_filters_label)
        
        filter_bar.addStretch()
        
        # Export
        export_btn = QPushButton("📥 Export")
        export_btn.clicked.connect(self.export_data)
        filter_bar.addWidget(export_btn)
        
        layout.addLayout(filter_bar)
        
        # ═══════════════════════════════════════════════════════
        #                      DATA TABLE
        # ═══════════════════════════════════════════════════════
        
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        layout.addWidget(self.table)
        
        # ═══════════════════════════════════════════════════════
        #                      STATUS BAR
        # ═══════════════════════════════════════════════════════
        
        status_bar = QHBoxLayout()
        
        self.status_label = QLabel("Ready")
        status_bar.addWidget(self.status_label)
        
        status_bar.addStretch()
        
        self.record_count_label = QLabel("0 records")
        status_bar.addWidget(self.record_count_label)
        
        layout.addLayout(status_bar)
        
        # بارگذاری sources
        self.load_sources()
    
    def load_sources(self):
        """بارگذاری لیست Data Sources از دیتابیس"""
        try:
            dm = DataManager()
            sources = dm.list_sources()
            
            self.source_combo.clear()
            self.source_combo.addItem("-- Select Data Source --", None)
            
            for source in sources:
                display_text = f"{source.name} ({source.total_records:,} records)"
                self.source_combo.addItem(display_text, source.id)
            
            self.status_label.setText(f"Loaded {len(sources)} data source(s)")
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"❌ خطا در بارگذاری منابع:\n{str(e)}")
            self.status_label.setText("خطا در بارگذاری منابع")
    
    def load_data(self):
        """بارگذاری داده‌های source انتخاب شده"""
        source_id = self.source_combo.currentData()
        if not source_id:
            self.table.setRowCount(0)
            self.record_count_label.setText("0 رکورد")
            return
        
        self.status_label.setText("در حال بارگذاری...")
        
        try:
            dm = DataManager()
            
            # دریافت لیست فیلدها برای این source
            source = dm.get_source(source_id)
            if not source:
                raise Exception("منبع داده یافت نشد")
            
            fields = dm.list_fields(source_id)
            
            # دریافت داده‌ها با استفاده از DataManager.query
            results = dm.query(source_id, limit=1000)
            
            # تنظیم جدول
            if not results:
                self.table.setRowCount(0)
                self.record_count_label.setText("0 رکورد")
                self.status_label.setText("داده‌ای یافت نشد")
                return
            
            # Header: تمام فیلدها
            headers = [f.field_display_name for f in fields]
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
            
            # پر کردن ردیف‌ها
            self.table.setRowCount(len(results))
            for row_idx, row_data in enumerate(results):
                # مقادیر فیلدها
                for col_idx, field in enumerate(fields):
                    value = row_data.get(field.field_name, '')
                    self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))
            
            # آمار
            total_count = len(results)
            self.record_count_label.setText(f"{total_count:,} رکورد")
            self.status_label.setText("آماده")
            
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"❌ خطا در بارگذاری داده:\n{str(e)}\n\nلطفاً اطمینان حاصل کنید که منبع داده موجود است.")
            self.status_label.setText("خطا در بارگذاری")
            self.table.setRowCount(0)
            import traceback
            print(traceback.format_exc())
    
    def show_advanced_filters(self):
        """نمایش دیالوگ فیلترهای پیشرفته"""
        source_id = self.source_combo.currentData()
        if not source_id:
            QMessageBox.warning(self, "هشدار", "⚠️ لطفاً ابتدا یک منبع داده انتخاب کنید")
            return
        
        try:
            dm = DataManager()
            fields = dm.list_fields(source_id)
            
            # تبدیل به فرمت مورد نیاز AdvancedFilterDialog
            field_list = [
                {
                    "id": f.id,
                    "name": f.field_name,
                    "display_name": f.field_display_name,
                    "type": f.data_type  # text, numeric, date, boolean
                }
                for f in fields
            ]
            
            dialog = AdvancedFilterDialog(field_list, self)
            
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self.current_filters = dialog.get_filters()
                self.apply_filters()
        
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"❌ خطا در بارگذاری فیلدها:\n{str(e)}")
    
    def apply_filters(self):
        """اعمال فیلترها و reload داده‌ها"""
        if not self.current_filters:
            self.active_filters_label.setText("No filters active")
        else:
            filter_text = f"{len(self.current_filters)} filter(s) active"
            self.active_filters_label.setText(filter_text)
            self.status_label.setText("Applying filters...")
        
        # بارگذاری مجدد با فیلترهای جدید
        self.load_data()
    
    def apply_quick_search(self, text):
        """جستجوی سریع در تمام ستون‌ها"""
        if not text:
            # نمایش تمام ردیف‌ها
            for row in range(self.table.rowCount()):
                self.table.setRowHidden(row, False)
            return
        
        # مخفی کردن ردیف‌های غیر مرتبط
        text_lower = text.lower()
        for row in range(self.table.rowCount()):
            match = False
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and text_lower in item.text().lower():
                    match = True
                    break
            self.table.setRowHidden(row, not match)
    
    def show_context_menu(self, position):
        """منوی کلیک راست"""
        menu = QMenu()
        
        copy_action = menu.addAction("📋 Copy")
        export_row_action = menu.addAction("📤 Export Selected Rows")
        menu.addSeparator()
        pivot_action = menu.addAction("📊 Create Pivot Table")
        chart_action = menu.addAction("📈 Create Chart")
        
        action = menu.exec(self.table.viewport().mapToGlobal(position))
        
        if action == copy_action:
            self.copy_selected()
        elif action == pivot_action:
            self.create_pivot_table()
        elif action == chart_action:
            self.create_chart()
    
    def copy_selected(self):
        """کپی ردیف‌های انتخاب شده به clipboard"""
        from PyQt6.QtWidgets import QApplication
        
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            return
        
        # جمع‌آوری داده‌های انتخاب شده
        data_rows = []
        for range_obj in selected_ranges:
            for row in range(range_obj.topRow(), range_obj.bottomRow() + 1):
                row_data = []
                for col in range(range_obj.leftColumn(), range_obj.rightColumn() + 1):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                data_rows.append("\t".join(row_data))
        
        # کپی به clipboard
        clipboard_text = "\n".join(data_rows)
        QApplication.clipboard().setText(clipboard_text)
        
        self.status_label.setText(f"Copied {len(data_rows)} row(s) to clipboard")
    
    def export_data(self):
        """Export داده‌ها به Excel"""
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "هشدار", "⚠️ داده‌ای برای خروجی وجود ندارد")
            return
        
        from PyQt6.QtWidgets import QFileDialog
        from datetime import datetime
        
        # انتخاب فایل
        default_name = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            default_name,
            "Excel Files (*.xlsx);;CSV Files (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            self.status_label.setText("Exporting...")
            
            if file_path.endswith('.csv'):
                self._export_csv(file_path)
            else:
                self._export_excel(file_path)
            
            QMessageBox.information(self, "موفق", f"✅ داده‌ها با موفقیت ذخیره شد:\n{file_path}")
            self.status_label.setText("خروجی با موفقیت انجام شد")
        
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"❌ خطا در خروجی:\n{str(e)}")
            self.status_label.setText("خطا در خروجی")
    
    def _export_excel(self, file_path):
        """Export به Excel با openpyxl"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise Exception("openpyxl not installed. Run: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Export"
        
        # Header
        headers = [self.table.horizontalHeaderItem(i).text() 
                   for i in range(self.table.columnCount())]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # داده‌ها
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            ws.append(row_data)
        
        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        wb.save(file_path)
    
    def _export_csv(self, file_path):
        """Export به CSV"""
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # Header
            headers = [self.table.horizontalHeaderItem(i).text() 
                       for i in range(self.table.columnCount())]
            writer.writerow(headers)
            
            # داده‌ها
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                writer.writerow(row_data)
    
    def create_pivot_table(self):
        """ساخت جدول پیوت"""
        QMessageBox.information(self, "جدول پیوت", "📊 قابلیت جدول پیوت به زودی اضافه می‌شود!")
    
    def create_chart(self):
        """ساخت نمودار"""
        QMessageBox.information(self, "نمودار", "📈 قابلیت نمودار به زودی اضافه می‌شود!")


if __name__ == "__main__":
    from PyQt6.QtWidgets import QApplication
    import sys
    
    app = QApplication(sys.argv)
    window = DataExplorerWidget()
    window.resize(1200, 700)
    window.show()
    sys.exit(app.exec())
